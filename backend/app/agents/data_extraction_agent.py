"""
Data Extraction Agent for Facultative Reinsurance System

This module provides comprehensive data extraction capabilities for reinsurance documents:
- Extracts structured data from OCR results
- Processes email attachments and content
- Generates Excel reports matching required format
- Uses Hugging Face models for entity extraction and classification
- Handles reinsurance-specific data fields and validation
"""

from __future__ import annotations

import logging
import os
import re
import json
from datetime import datetime
from typing import Dict, List, Optional, Any, Union, Tuple
from pathlib import Path
import pandas as pd
import numpy as np

# Hugging Face transformers
try:
    from transformers import (
        AutoTokenizer, AutoModelForTokenClassification,
        AutoModelForSequenceClassification, pipeline,
        AutoModel
    )
    TRANSFORMERS_AVAILABLE = True
except ImportError:
    TRANSFORMERS_AVAILABLE = False

# Pydantic models
from pydantic import BaseModel, Field, validator
from decimal import Decimal

# Import OCR agent components
from .ocr_agent import (
    ocr_agent, OCRResult, EmailContent, ExcelData, 
    WordDocumentData, PowerPointData, AttachmentData
)

# Import RI Slip identification components
from .ri_slip_identifier import (
    ri_slip_identifier, RISlipIdentificationResult, 
    EmailProcessingPlan, DocumentClassificationResult
)

# Import validation components
from ..services.validation import ValidationResult

# Configure logging
logger = logging.getLogger(__name__)


class FinancialData(BaseModel):
    """Financial data extracted from documents"""
    sum_insured: Optional[Decimal] = None
    premium: Optional[Decimal] = None
    deductible: Optional[Decimal] = None
    limit: Optional[Decimal] = None
    currency: Optional[str] = None
    policy_period: Optional[str] = None
    coverage_type: Optional[str] = None


class RiskData(BaseModel):
    """Risk-related data extracted from documents"""
    risk_type: Optional[str] = None
    location: Optional[str] = None
    country: Optional[str] = None
    industry: Optional[str] = None
    occupancy: Optional[str] = None
    construction_type: Optional[str] = None
    year_built: Optional[int] = None
    risk_description: Optional[str] = None


class PartyData(BaseModel):
    """Party information (insured, broker, etc.)"""
    name: Optional[str] = None
    address: Optional[str] = None
    contact_person: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    role: Optional[str] = None  # insured, broker, reinsurer, etc.


class ExtractedData(BaseModel):
    """Complete extracted data structure with Analysis document support"""
    # Legacy fields for backward compatibility
    document_type: Optional[str] = None
    reference_number: Optional[str] = None
    date: Optional[datetime] = None
    
    # Financial information (legacy)
    financial: FinancialData = Field(default_factory=FinancialData)
    
    # Risk information (legacy)
    risk: RiskData = Field(default_factory=RiskData)
    
    # Parties involved (legacy)
    parties: List[PartyData] = Field(default_factory=list)
    
    # Reinsurance-specific parties (legacy)
    broker: Optional[str] = None
    insured: Optional[str] = None  # End policyholder
    reinsured: Optional[str] = None  # Same as cedant/insurer - the ceding company
    cedant: Optional[str] = None  # Alternative term for reinsured
    insurer: Optional[str] = None  # Alternative term for reinsured
    reinsurer: Optional[str] = None  # Company providing reinsurance
    
    # Business details (legacy)
    line_of_business: Optional[str] = None
    treaty_type: Optional[str] = None
    attachment_point: Optional[Decimal] = None
    
    # Additional legacy fields
    underwriter: Optional[str] = None
    account_handler: Optional[str] = None
    renewal_date: Optional[datetime] = None
    inception_date: Optional[datetime] = None
    expiry_date: Optional[datetime] = None
    
    # NEW: Analysis document data (23 critical fields)
    analysis_data: 'AnalysisDocumentData' = Field(default_factory=lambda: AnalysisDocumentData())
    
    # Metadata
    confidence_score: float = 0.0
    extraction_method: str = "hybrid"
    processing_notes: List[str] = Field(default_factory=list)
    data_completeness: Dict[str, bool] = Field(default_factory=dict)
    
    def migrate_to_analysis_format(self) -> None:
        """Migrate legacy fields to Analysis document format"""
        # Map legacy fields to analysis document fields
        if self.reference_number and not self.analysis_data.reference_number:
            self.analysis_data.reference_number = self.reference_number
        
        if self.insured and not self.analysis_data.insured_name:
            self.analysis_data.insured_name = self.insured
        
        if self.reinsured and not self.analysis_data.cedant_reinsured:
            self.analysis_data.cedant_reinsured = self.reinsured
        elif self.cedant and not self.analysis_data.cedant_reinsured:
            self.analysis_data.cedant_reinsured = self.cedant
        
        if self.broker and not self.analysis_data.broker_name:
            self.analysis_data.broker_name = self.broker
        
        if self.financial.sum_insured and not self.analysis_data.total_sums_insured:
            self.analysis_data.total_sums_insured = self.financial.sum_insured
        
        if self.financial.currency and not self.analysis_data.currency:
            self.analysis_data.currency = self.financial.currency
        
        if self.financial.deductible and not self.analysis_data.excess_retention:
            self.analysis_data.excess_retention = self.financial.deductible
        
        if self.date and not self.analysis_data.date_received:
            self.analysis_data.date_received = self.date
        
        if self.risk.location and not self.analysis_data.situation_of_risk:
            self.analysis_data.situation_of_risk = self.risk.location
        
        if self.risk.industry and not self.analysis_data.occupation_of_insured:
            self.analysis_data.occupation_of_insured = self.risk.industry
        
        # Update confidence and completeness
        self.analysis_data.confidence_score = self.confidence_score
        self.analysis_data.processing_notes = self.processing_notes.copy()
    
    def to_analysis_excel_row(self) -> Dict[str, Any]:
        """Convert to Analysis document Excel row format"""
        self.migrate_to_analysis_format()
        
        row_data = {}
        field_mapping = {
            'Reference Number': 'reference_number',
            'Date Received': 'date_received',
            'Insured': 'insured_name',
            'Cedant/Reinsured': 'cedant_reinsured',
            'Broker': 'broker_name',
            'Perils Covered': 'perils_covered',
            'Geographical Limit': 'geographical_limit',
            'Situation of Risk/Voyage': 'situation_of_risk',
            'Occupation of Insured': 'occupation_of_insured',
            'Main Activities': 'main_activities',
            'Total Sums Insured': 'total_sums_insured',
            'Currency': 'currency',
            'Excess/Retention': 'excess_retention',
            'Premium Rates (%)': 'premium_rates',
            'Period of Insurance': 'period_of_insurance',
            'PML %': 'pml_percentage',
            'CAT Exposure': 'cat_exposure',
            'Reinsurance Deductions': 'reinsurance_deductions',
            'Claims Experience (3 years)': 'claims_experience_3_years',
            'Share offered %': 'share_offered_percentage',
            'Surveyor\'s Report': 'surveyors_report',
            'Climate Change Risk': 'climate_change_risk',
            'ESG Risk Assessment': 'esg_risk_assessment'
        }
        
        for excel_col, field_name in field_mapping.items():
            value = getattr(self.analysis_data, field_name, None)
            
            # Format values appropriately for Excel
            if isinstance(value, Decimal):
                row_data[excel_col] = float(value)
            elif isinstance(value, datetime):
                row_data[excel_col] = value.strftime('%Y-%m-%d')
            else:
                row_data[excel_col] = value
        
        # Add metadata columns
        row_data['Confidence Score'] = self.analysis_data.confidence_score
        row_data['Data Completeness %'] = self.analysis_data.calculate_completeness_score() * 100
        row_data['Processing Notes'] = '; '.join(self.analysis_data.processing_notes)
        row_data['Source Documents'] = '; '.join(self.analysis_data.source_documents)
        
        return row_data


# Analysis Document Data Model (imported from schemas)
class AnalysisDocumentData(BaseModel):
    """
    Analysis document data structure with all 23 critical fields
    Based on standard reinsurance analysis working sheet format
    """
    # Basic Information (Fields 1-5)
    reference_number: Optional[str] = Field(None, max_length=50, description="Unique reference/policy number")
    date_received: Optional[datetime] = Field(None, description="Date document was received")
    insured_name: Optional[str] = Field(None, max_length=200, description="Name of the insured party")
    cedant_reinsured: Optional[str] = Field(None, max_length=200, description="Ceding company/reinsured")
    broker_name: Optional[str] = Field(None, max_length=200, description="Broker/intermediary name")
    
    # Coverage Details (Fields 6-10)
    perils_covered: Optional[str] = Field(None, max_length=500, description="Perils/risks covered")
    geographical_limit: Optional[str] = Field(None, max_length=300, description="Geographic coverage limits")
    situation_of_risk: Optional[str] = Field(None, max_length=500, description="Risk location/voyage details")
    occupation_of_insured: Optional[str] = Field(None, max_length=200, description="Insured's business/occupation")
    main_activities: Optional[str] = Field(None, max_length=500, description="Main business activities")
    
    # Financial Information (Fields 11-15)
    total_sums_insured: Optional[Decimal] = Field(None, ge=0, description="Total sum insured amount")
    currency: Optional[str] = Field(None, max_length=3, description="Currency code (USD, EUR, etc.)")
    excess_retention: Optional[Decimal] = Field(None, ge=0, description="Excess/retention amount")
    premium_rates: Optional[Decimal] = Field(None, ge=0, le=100, description="Premium rate percentage")
    period_of_insurance: Optional[str] = Field(None, max_length=100, description="Insurance period")
    
    # Risk Assessment (Fields 16-20)
    pml_percentage: Optional[Decimal] = Field(None, ge=0, le=100, description="Probable Maximum Loss percentage")
    cat_exposure: Optional[str] = Field(None, max_length=300, description="Catastrophe exposure details")
    reinsurance_deductions: Optional[Decimal] = Field(None, ge=0, description="Reinsurance deductions")
    claims_experience_3_years: Optional[str] = Field(None, max_length=1000, description="3-year claims history")
    share_offered_percentage: Optional[Decimal] = Field(None, ge=0, le=100, description="Share offered percentage")
    
    # Additional Information (Fields 21-23)
    surveyors_report: Optional[str] = Field(None, max_length=500, description="Surveyor's report details")
    climate_change_risk: Optional[str] = Field(None, max_length=500, description="Climate change risk assessment")
    esg_risk_assessment: Optional[str] = Field(None, max_length=500, description="ESG risk factors")
    
    # Metadata and Processing Information
    confidence_score: float = Field(0.0, ge=0.0, le=1.0, description="Overall extraction confidence")
    field_confidence_scores: Dict[str, float] = Field(default_factory=dict, description="Per-field confidence scores")
    data_completeness: Dict[str, bool] = Field(default_factory=dict, description="Field completeness tracking")
    processing_notes: List[str] = Field(default_factory=list, description="Processing notes and warnings")
    extraction_method: str = Field("hybrid", description="Extraction method used")
    source_documents: List[str] = Field(default_factory=list, description="Source document references")
    
    @validator('currency')
    def validate_currency(cls, v):
        """Validate currency code format"""
        if v and len(v) != 3:
            raise ValueError('Currency code must be 3 characters')
        return v.upper() if v else v
    
    @validator('premium_rates', 'pml_percentage', 'share_offered_percentage')
    def validate_percentages(cls, v):
        """Validate percentage fields are within valid range"""
        if v is not None and (v < 0 or v > 100):
            raise ValueError('Percentage must be between 0 and 100')
        return v
    
    @validator('total_sums_insured', 'excess_retention', 'reinsurance_deductions')
    def validate_financial_amounts(cls, v):
        """Validate financial amounts are non-negative"""
        if v is not None and v < 0:
            raise ValueError('Financial amounts must be non-negative')
        return v
    
    def calculate_completeness_score(self) -> float:
        """Calculate data completeness score based on critical fields"""
        critical_fields = [
            'reference_number', 'insured_name', 'cedant_reinsured', 'broker_name',
            'perils_covered', 'total_sums_insured', 'currency', 'period_of_insurance',
            'pml_percentage', 'share_offered_percentage'
        ]
        
        completed_fields = sum(1 for field in critical_fields if getattr(self, field) is not None)
        return completed_fields / len(critical_fields) if critical_fields else 0.0
    
    def get_missing_critical_fields(self) -> List[str]:
        """Get list of missing critical fields"""
        critical_fields = [
            'reference_number', 'insured_name', 'cedant_reinsured', 'broker_name',
            'perils_covered', 'total_sums_insured', 'currency', 'period_of_insurance',
            'pml_percentage', 'share_offered_percentage'
        ]
        
        return [field for field in critical_fields if getattr(self, field) is None]


class DataExtractionAgent:
    """
    Advanced data extraction agent for reinsurance documents
    
    Combines OCR results with NLP models to extract structured data
    and generate Excel reports in the required format.
    """
    
    def __init__(self):
        """Initialize the data extraction agent"""
        self.models = {}
        # Cache dir for HF models
        self.hf_cache_dir = os.environ.get("HF_HOME", "/app/.cache/huggingface")
        self._initialize_models()
        
        # Reinsurance-specific patterns and keywords
        self._initialize_patterns()
        
        # Excel template structure
        self._initialize_excel_template()
    
    def _initialize_models(self):
        """Initialize Hugging Face models for data extraction"""
        if not TRANSFORMERS_AVAILABLE:
            logger.warning("Transformers not available, using pattern-based extraction only")
            return
        
        try:
            # Preload models/tokenizers with cache_dir, then build pipelines without cache_dir
            # NER model
            ner_tokenizer = AutoTokenizer.from_pretrained(
                "dbmdz/bert-large-cased-finetuned-conll03-english", cache_dir=str(self.hf_cache_dir)
            )
            ner_model = AutoModelForTokenClassification.from_pretrained(
                "dbmdz/bert-large-cased-finetuned-conll03-english", cache_dir=str(self.hf_cache_dir)
            )
            self.models['ner'] = pipeline(
                "ner",
                model=ner_model,
                tokenizer=ner_tokenizer,
                aggregation_strategy="simple"
            )

            # Financial NER (FinBERT as placeholder pipeline)
            finbert_tokenizer = AutoTokenizer.from_pretrained(
                "ProsusAI/finbert", cache_dir=str(self.hf_cache_dir)
            )
            finbert_model = AutoModelForTokenClassification.from_pretrained(
                "ProsusAI/finbert", cache_dir=str(self.hf_cache_dir)
            )
            self.models['finbert'] = pipeline(
                "ner",
                model=finbert_model,
                tokenizer=finbert_tokenizer,
                aggregation_strategy="simple"
            )

            # Zero-shot classifier
            zsl_tokenizer = AutoTokenizer.from_pretrained(
                "facebook/bart-large-mnli", cache_dir=str(self.hf_cache_dir)
            )
            zsl_model = AutoModelForSequenceClassification.from_pretrained(
                "facebook/bart-large-mnli", cache_dir=str(self.hf_cache_dir)
            )
            self.models['classifier'] = pipeline(
                "zero-shot-classification",
                model=zsl_model,
                tokenizer=zsl_tokenizer
            )

            # Sentiment (using FinBERT sequence classification variant if applicable)
            sent_tokenizer = AutoTokenizer.from_pretrained(
                "ProsusAI/finbert", cache_dir=str(self.hf_cache_dir)
            )
            sent_model = AutoModelForSequenceClassification.from_pretrained(
                "ProsusAI/finbert", cache_dir=str(self.hf_cache_dir)
            )
            self.models['sentiment'] = pipeline(
                "sentiment-analysis",
                model=sent_model,
                tokenizer=sent_tokenizer
            )

            logger.info("Hugging Face models initialized successfully")

        except Exception as e:
            logger.error(f"Failed to initialize models: {str(e)}")
            self.models = {}
    
    def _initialize_patterns(self):
        """Initialize regex patterns for data extraction"""
        self.patterns = {
            # Financial amounts
            'currency_amount': re.compile(
                r'(?:USD|EUR|GBP|CAD|AUD|JPY|CHF|SEK|NOK|DKK|INR|CNY|BRL|MXN|ZAR|SGD|HKD|NZD|THB|MYR|IDR|PHP|VND|KRW|TWD|RUB|TRY|PLN|CZK|HUF|RON|BGN|HRK|RSD|BAM|MKD|ALL|EUR|USD|GBP|CAD|AUD|JPY|CHF|SEK|NOK|DKK|INR|CNY|BRL|MXN|ZAR|SGD|HKD|NZD|THB|MYR|IDR|PHP|VND|KRW|TWD|RUB|TRY|PLN|CZK|HUF|RON|BGN|HRK|RSD|BAM|MKD|ALL)\s*[\d,]+(?:\.\d{2})?',
                re.IGNORECASE
            ),
            
            # Policy numbers and references
            'policy_number': re.compile(
                r'(?:policy|pol|ref|reference|no|number|#)\s*:?\s*([A-Z0-9\-/]+)',
                re.IGNORECASE
            ),
            
            # Dates
            'date': re.compile(
                r'\b(?:\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{2,4}[/-]\d{1,2}[/-]\d{1,2}|\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{2,4})',
                re.IGNORECASE
            ),
            
            # Email addresses
            'email': re.compile(
                r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
            ),
            
            # Phone numbers
            'phone': re.compile(
                r'(?:\+?1[-.\s]?)?\(?[0-9]{3}\)?[-.\s]?[0-9]{3}[-.\s]?[0-9]{4}'
            ),
            
            # Percentages
            'percentage': re.compile(
                r'\b\d+(?:\.\d+)?%\b'
            ),
            
            # Geographic locations
            'location': re.compile(
                r'\b(?:located|situated|at|in)\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)',
                re.IGNORECASE
            ),
            
            # NEW: Specialized reinsurance field patterns (Task 5.3)
            
            # Basic field patterns
            'insured_pattern': re.compile(
                r'(?:insured|assured|policyholder|risk)\s*:?\s*([A-Z][A-Za-z\s&.,()-]+(?:Ltd|Inc|Corp|Company|Insurance|Group|plc|LLC|AG|SA|SE|NV|BV|Limited|Corporation)?)',
                re.IGNORECASE
            ),
            
            'cedant_pattern': re.compile(
                r'(?:cedant|ceding company|reinsured|original insurer|insurer)\s*:?\s*([A-Z][A-Za-z\s&.,()-]+(?:Ltd|Inc|Corp|Company|Insurance|Group|plc|LLC|AG|SA|SE|NV|BV|Limited|Corporation)?)',
                re.IGNORECASE
            ),
            
            'broker_pattern': re.compile(
                r'(?:broker|intermediary|placing broker|reinsurance broker|agent)\s*:?\s*([A-Z][A-Za-z\s&.,()-]+(?:Ltd|Inc|Corp|Company|Insurance|Group|plc|LLC|AG|SA|SE|NV|BV|Limited|Corporation|Brokers)?)',
                re.IGNORECASE
            ),
            
            'perils_pattern': re.compile(
                r'(?:perils?\s+covered|coverage|risks?\s+covered|insured\s+perils?)\s*:?\s*([A-Za-z\s,&/()-]+(?:fire|flood|earthquake|wind|storm|theft|explosion|terrorism|war|damage|loss|liability|accident|injury|death|business\s+interruption|machinery\s+breakdown|cyber|data\s+breach)[A-Za-z\s,&/()-]*)',
                re.IGNORECASE | re.DOTALL
            ),
            
            # Geographic and risk patterns
            'geographical_limit_pattern': re.compile(
                r'(?:geographical?\s+limit|territory|jurisdiction|coverage\s+area|geographic\s+scope)\s*:?\s*([A-Za-z\s,&/()-]+(?:worldwide|global|excluding|including|limited\s+to|USA|Europe|Asia|Africa|America|Australia|country|state|province|city)[A-Za-z\s,&/()-]*)',
                re.IGNORECASE | re.DOTALL
            ),
            
            'situation_of_risk_pattern': re.compile(
                r'(?:situation\s+of\s+risk|risk\s+location|voyage|premises|address|location|site|facility)\s*:?\s*([A-Za-z0-9\s,&/().-]+(?:street|road|avenue|boulevard|drive|lane|way|building|floor|suite|city|state|province|country|port|terminal|warehouse|factory|plant|office)[A-Za-z0-9\s,&/().-]*)',
                re.IGNORECASE | re.DOTALL
            ),
            
            # Business patterns
            'occupation_pattern': re.compile(
                r'(?:occupation\s+of\s+insured|business|industry|sector|trade|profession)\s*:?\s*([A-Za-z\s&/()-]+(?:manufacturing|retail|wholesale|construction|healthcare|technology|finance|education|hospitality|transportation|energy|mining|agriculture|consulting|government)[A-Za-z\s&/()-]*)',
                re.IGNORECASE | re.DOTALL
            ),
            
            'main_activities_pattern': re.compile(
                r'(?:main\s+activities|business\s+activities|operations|activities|description\s+of\s+business)\s*:?\s*([A-Za-z0-9\s,&/().-]+(?:production|manufacturing|sales|distribution|services|consulting|development|research|processing|trading|import|export|retail|wholesale)[A-Za-z0-9\s,&/().-]*)',
                re.IGNORECASE | re.DOTALL
            ),
            
            # Financial patterns
            'total_sum_insured_pattern': re.compile(
                r'(?:total\s+sum\s+insured|tsi|sum\s+insured|insured\s+value|coverage\s+limit|policy\s+limit|limit)\s*:?\s*([A-Z]{3})?\s*[\$€£¥]?\s*([\d,]+(?:\.\d{2})?)\s*(?:million|m|thousand|k|billion|b)?',
                re.IGNORECASE
            ),
            
            'excess_retention_pattern': re.compile(
                r'(?:excess|deductible|retention|self\s+insured\s+retention|sir|franchise|first\s+loss|attachment\s+point)\s*:?\s*([A-Z]{3})?\s*[\$€£¥]?\s*([\d,]+(?:\.\d{2})?)\s*(?:million|m|thousand|k|billion|b)?',
                re.IGNORECASE
            ),
            
            'premium_rates_pattern': re.compile(
                r'(?:premium\s+rate|rate|pricing|cost|premium\s+%|rate\s+%|premium\s+per\s+mille|rate\s+per\s+thousand|basis\s+points)\s*:?\s*([\d.]+)\s*(?:%|percent|per\s+cent|bps|basis\s+points|‰|per\s+mille)?',
                re.IGNORECASE
            ),
            
            'period_of_insurance_pattern': re.compile(
                r'(?:period\s+of\s+insurance|policy\s+period|coverage\s+period|term|effective\s+date|inception\s+date|expiry\s+date|renewal\s+date)\s*:?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}(?:\s+to\s+\d{1,2}[/-]\d{1,2}[/-]\d{2,4})?|\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{2,4}(?:\s+to\s+\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{2,4})?|12\s+months?|1\s+year|24\s+months?|2\s+years?|36\s+months?|3\s+years?)',
                re.IGNORECASE
            ),
            
            # NEW: Advanced reinsurance-specific field patterns (Task 5.4)
            
            # PML % and CAT Exposure patterns
            'pml_percentage_pattern': re.compile(
                r'(?:pml|probable\s+maximum\s+loss|maximum\s+probable\s+loss|mpl)\s*:?\s*([\d.]+)\s*(?:%|percent|per\s+cent)?',
                re.IGNORECASE
            ),
            
            'cat_exposure_pattern': re.compile(
                r'(?:cat\s+exposure|catastrophe\s+exposure|natural\s+catastrophe|nat\s+cat|earthquake|flood|hurricane|windstorm|typhoon|tsunami|volcanic|wildfire|hail|tornado|cyclone)\s*:?\s*([A-Za-z0-9\s,&/().-]+(?:zone|region|area|exposure|risk|peril|hazard|model|rms|air|eqecat|karen\s+clark)[A-Za-z0-9\s,&/().-]*)',
                re.IGNORECASE | re.DOTALL
            ),
            
            # Period of Insurance and Reinsurance Deductions patterns
            'period_of_insurance_detailed_pattern': re.compile(
                r'(?:period\s+of\s+insurance|policy\s+period|coverage\s+period|insurance\s+term|effective\s+period|risk\s+period)\s*:?\s*(?:from\s+)?(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{2,4})\s*(?:to\s+)?(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{2,4})?|(?:12\s+months?|1\s+year|24\s+months?|2\s+years?|36\s+months?|3\s+years?)',
                re.IGNORECASE
            ),
            
            'reinsurance_deductions_pattern': re.compile(
                r'(?:reinsurance\s+deductions?|ri\s+deductions?|deductions?\s+from\s+reinsurance|reinsurance\s+commission|brokerage|override|profit\s+commission|sliding\s+scale|no\s+claims\s+bonus|ncb)\s*:?\s*([A-Z]{3})?\s*[\$€£¥]?\s*([\d,]+(?:\.\d{2})?)\s*(?:million|m|thousand|k|billion|b)?|(?:[\d.]+)\s*(?:%|percent|per\s+cent)',
                re.IGNORECASE
            ),
            
            # Claims Experience and Share offered patterns
            'claims_experience_pattern': re.compile(
                r'(?:claims?\s+experience|loss\s+experience|claims?\s+history|loss\s+history|claims?\s+record|loss\s+record)\s*(?:(?:for\s+)?(?:the\s+)?(?:last\s+|past\s+)?(?:3\s+years?|three\s+years?|36\s+months?))\s*:?\s*([A-Za-z0-9\s,&/().$€£¥-]+(?:nil|none|no\s+claims?|no\s+losses?|clean|excellent|good|poor|bad|claims?|losses?|paid|outstanding|incurred|reported|settled)[A-Za-z0-9\s,&/().$€£¥-]*)',
                re.IGNORECASE | re.DOTALL
            ),
            
            'share_offered_percentage_pattern': re.compile(
                r'(?:share\s+offered|percentage\s+offered|% offered|quota\s+share|participation|line\s+size|capacity\s+offered|reinsurance\s+share|ri\s+share)\s*:?\s*([\d.]+)\s*(?:%|percent|per\s+cent)',
                re.IGNORECASE
            ),
            
            # Surveyor's report detection and attachment linking
            'surveyors_report_pattern': re.compile(
                r'(?:surveyor\'?s?\s+report|survey\s+report|inspection\s+report|risk\s+survey|engineering\s+survey|loss\s+control\s+survey|pre-risk\s+survey|condition\s+survey)\s*:?\s*([A-Za-z0-9\s,&/().-]+(?:attached|available|provided|enclosed|separate|pending|required|recommended|completed|dated|issued|prepared)[A-Za-z0-9\s,&/().-]*)',
                re.IGNORECASE | re.DOTALL
            ),
            
            'surveyors_report_attachment_pattern': re.compile(
                r'(?:attachment|attached|file|document|report)\s*:?\s*([A-Za-z0-9_.-]+\.(?:pdf|doc|docx|xls|xlsx|jpg|jpeg|png|tiff?))',
                re.IGNORECASE
            ),
            
            # Climate Change and ESG Risk Assessment patterns
            'climate_change_risk_pattern': re.compile(
                r'(?:climate\s+change|climate\s+risk|global\s+warming|carbon\s+footprint|greenhouse\s+gas|emissions|sustainability|environmental\s+impact|physical\s+climate\s+risk|transition\s+risk|stranded\s+assets|carbon\s+pricing|renewable\s+energy|green\s+technology|climate\s+adaptation|climate\s+resilience)\s*:?\s*([A-Za-z0-9\s,&/().-]+(?:low|medium|high|significant|minimal|moderate|severe|critical|assessed|evaluated|considered|impact|risk|exposure|mitigation|adaptation|strategy|policy|commitment|target|goal)[A-Za-z0-9\s,&/().-]*)',
                re.IGNORECASE | re.DOTALL
            ),
            
            'esg_risk_assessment_pattern': re.compile(
                r'(?:esg|environmental\s+social\s+governance|environmental\s+social\s+and\s+governance|sustainability|corporate\s+responsibility|csr|social\s+responsibility|governance\s+risk|environmental\s+risk|social\s+risk|ethical\s+risk|reputation\s+risk|stakeholder\s+risk|regulatory\s+risk|compliance\s+risk)\s*:?\s*([A-Za-z0-9\s,&/().-]+(?:low|medium|high|significant|minimal|moderate|severe|critical|assessed|evaluated|considered|impact|risk|exposure|rating|score|framework|policy|commitment|target|goal|compliance|non-compliance|violation|breach)[A-Za-z0-9\s,&/().-]*)',
                re.IGNORECASE | re.DOTALL
            )
        }
        
        # Reinsurance-specific keywords
        self.keywords = {
            'document_types': [
                'facultative certificate', 'treaty slip', 'placement slip',
                'risk placement', 'reinsurance slip', 'cover note',
                'binding authority', 'line slip', 'open cover', 'ri slip',
                'fac slip', 'facultative slip', 'treaty certificate'
            ],
            
            'coverage_types': [
                'property', 'casualty', 'marine', 'aviation', 'energy',
                'cyber', 'political risk', 'credit', 'surety', 'life',
                'health', 'accident', 'motor', 'fire', 'flood', 'earthquake',
                'cat', 'catastrophe', 'wind', 'storm', 'terrorism'
            ],
            
            'treaty_types': [
                'quota share', 'surplus', 'excess of loss', 'stop loss',
                'aggregate excess', 'catastrophe', 'per risk', 'per occurrence',
                'working excess', 'clash', 'aggregate', 'cat xl', 'working xl'
            ],
            
            'currencies': [
                'USD', 'EUR', 'GBP', 'CAD', 'AUD', 'JPY', 'CHF', 'SEK',
                'NOK', 'DKK', 'INR', 'CNY', 'BRL', 'MXN', 'ZAR', 'SGD',
                'HKD', 'NZD', 'THB', 'MYR', 'KRW', 'TWD', 'PLN', 'CZK'
            ],
            
            # Reinsurance party identification keywords
            'party_keywords': {
                'reinsured': ['reinsured', 'cedant', 'ceding company', 'insurer', 'original insurer'],
                'reinsurer': ['reinsurer', 'reinsurance company', 'accepting office'],
                'broker': ['broker', 'intermediary', 'placing broker', 'reinsurance broker'],
                'insured': ['insured', 'policyholder', 'assured', 'risk'],
                'underwriter': ['underwriter', 'lead underwriter', 'uw'],
                'account_handler': ['account handler', 'account manager', 'relationship manager']
            },
            
            # NEW: Specialized reinsurance field keywords (Task 5.3)
            'perils_keywords': [
                'fire', 'explosion', 'earthquake', 'flood', 'windstorm', 'hurricane', 'typhoon',
                'tornado', 'hail', 'lightning', 'theft', 'burglary', 'vandalism', 'malicious damage',
                'terrorism', 'war', 'civil commotion', 'riot', 'strike', 'impact', 'aircraft',
                'vehicle', 'smoke', 'water damage', 'leakage', 'burst pipes', 'sprinkler leakage',
                'subsidence', 'landslide', 'volcanic eruption', 'tsunami', 'storm surge',
                'business interruption', 'machinery breakdown', 'boiler explosion', 'cyber attack',
                'data breach', 'system failure', 'power outage', 'contamination', 'pollution'
            ],
            
            'geographical_keywords': [
                'worldwide', 'global', 'international', 'territory', 'jurisdiction', 'excluding',
                'including', 'limited to', 'restricted to', 'coverage area', 'geographic limit',
                'territorial limit', 'zone', 'region', 'country', 'state', 'province', 'city',
                'address', 'location', 'premises', 'site', 'facility', 'plant', 'building'
            ],
            
            'occupation_keywords': [
                'manufacturing', 'retail', 'wholesale', 'distribution', 'logistics', 'transportation',
                'construction', 'real estate', 'hospitality', 'restaurant', 'hotel', 'healthcare',
                'education', 'technology', 'software', 'telecommunications', 'energy', 'utilities',
                'mining', 'agriculture', 'forestry', 'fishing', 'banking', 'finance', 'insurance',
                'consulting', 'professional services', 'government', 'non-profit', 'entertainment',
                'media', 'publishing', 'pharmaceutical', 'chemical', 'automotive', 'aerospace',
                'defense', 'oil and gas', 'renewable energy', 'waste management', 'recycling'
            ],
            
            'financial_keywords': {
                'sum_insured': ['sum insured', 'total sum insured', 'tsi', 'limit', 'coverage limit', 
                               'policy limit', 'insured value', 'insured amount', 'coverage amount'],
                'excess_retention': ['excess', 'deductible', 'retention', 'self insured retention', 
                                   'sir', 'franchise', 'first loss', 'attachment point'],
                'premium_rates': ['premium rate', 'rate', 'pricing', 'cost', 'premium %', 'rate %',
                                'premium per mille', 'rate per thousand', 'basis points'],
                'period': ['period of insurance', 'policy period', 'coverage period', 'term',
                          'effective date', 'expiry date', 'inception', 'renewal date']
            },
            
            # NEW: Advanced reinsurance field keywords (Task 5.4)
            'pml_keywords': [
                'pml', 'probable maximum loss', 'maximum probable loss', 'mpl', 'worst case scenario',
                'maximum foreseeable loss', 'mfl', 'estimated maximum loss', 'eml', 'maximum possible loss',
                'catastrophe loss', 'single event loss', 'aggregate loss', 'occurrence loss'
            ],
            
            'cat_exposure_keywords': [
                'catastrophe exposure', 'cat exposure', 'natural catastrophe', 'nat cat', 'peril exposure',
                'earthquake', 'flood', 'hurricane', 'windstorm', 'typhoon', 'cyclone', 'tsunami',
                'volcanic eruption', 'wildfire', 'hailstorm', 'tornado', 'storm surge', 'landslide',
                'subsidence', 'drought', 'freeze', 'winter storm', 'severe weather', 'climate peril',
                'catastrophe zone', 'hazard zone', 'exposure zone', 'risk zone', 'peril zone',
                'rms model', 'air model', 'eqecat model', 'karen clark model', 'catastrophe model'
            ],
            
            'reinsurance_deductions_keywords': [
                'reinsurance deductions', 'ri deductions', 'deductions from reinsurance', 'reinsurance commission',
                'brokerage', 'override commission', 'profit commission', 'sliding scale commission',
                'no claims bonus', 'ncb', 'experience refund', 'loss participation', 'ceding commission',
                'acquisition costs', 'management fee', 'service fee', 'administration fee'
            ],
            
            'claims_experience_keywords': [
                'claims experience', 'loss experience', 'claims history', 'loss history', 'claims record',
                'loss record', 'loss ratio', 'claims ratio', 'frequency', 'severity', 'incurred losses',
                'paid losses', 'outstanding losses', 'reported losses', 'settled claims', 'open claims',
                'nil claims', 'no claims', 'clean record', 'excellent experience', 'poor experience',
                'adverse experience', 'favorable experience', 'loss development', 'claims development'
            ],
            
            'share_offered_keywords': [
                'share offered', 'percentage offered', 'quota share', 'participation', 'line size',
                'capacity offered', 'reinsurance share', 'ri share', 'retention', 'cession',
                'written line', 'signed line', 'order', 'subscription', 'participation percentage'
            ],
            
            'surveyors_report_keywords': [
                'surveyor report', 'survey report', 'inspection report', 'risk survey', 'engineering survey',
                'loss control survey', 'pre-risk survey', 'condition survey', 'risk assessment report',
                'technical survey', 'fire survey', 'security survey', 'housekeeping survey',
                'recommendations', 'risk improvements', 'risk mitigation', 'loss prevention',
                'attached', 'available', 'provided', 'enclosed', 'separate', 'pending', 'required'
            ],
            
            'climate_change_keywords': [
                'climate change', 'climate risk', 'global warming', 'carbon footprint', 'greenhouse gas',
                'emissions', 'sustainability', 'environmental impact', 'physical climate risk',
                'transition risk', 'stranded assets', 'carbon pricing', 'renewable energy',
                'green technology', 'climate adaptation', 'climate resilience', 'climate mitigation',
                'carbon neutral', 'net zero', 'decarbonization', 'climate scenario', 'tcfd'
            ],
            
            'esg_keywords': [
                'esg', 'environmental social governance', 'environmental social and governance',
                'sustainability', 'corporate responsibility', 'csr', 'social responsibility',
                'governance risk', 'environmental risk', 'social risk', 'ethical risk',
                'reputation risk', 'stakeholder risk', 'regulatory risk', 'compliance risk',
                'diversity', 'inclusion', 'human rights', 'labor practices', 'supply chain',
                'board composition', 'executive compensation', 'transparency', 'accountability',
                'anti-corruption', 'data privacy', 'cybersecurity', 'business ethics'
            ]
        }
    
    def _initialize_excel_template(self):
        """Initialize Excel template structure based on Analysis document format (23 critical fields)"""
        self.excel_template = {
            # Analysis document format with all 23 critical fields
            'columns': [
                # Basic Information (Fields 1-5)
                'Reference Number',
                'Date Received',
                'Insured',
                'Cedant/Reinsured',
                'Broker',
                
                # Coverage Details (Fields 6-10)
                'Perils Covered',
                'Geographical Limit',
                'Situation of Risk/Voyage',
                'Occupation of Insured',
                'Main Activities',
                
                # Financial Information (Fields 11-15)
                'Total Sums Insured',
                'Currency',
                'Excess/Retention',
                'Premium Rates (%)',
                'Period of Insurance',
                
                # Risk Assessment (Fields 16-20)
                'PML %',
                'CAT Exposure',
                'Reinsurance Deductions',
                'Claims Experience (3 years)',
                'Share offered %',
                
                # Additional Information (Fields 21-23)
                'Surveyor\'s Report',
                'Climate Change Risk',
                'ESG Risk Assessment',
                
                # Metadata columns
                'Confidence Score',
                'Data Completeness %',
                'Processing Notes',
                'Source Documents'
            ],
            
            # Fields that are commonly available from RI Slips
            'commonly_available': [
                'Reference Number', 'Insured', 'Cedant/Reinsured', 'Broker', 
                'Total Sums Insured', 'Currency', 'Perils Covered', 
                'Situation of Risk/Voyage', 'Period of Insurance'
            ],
            
            # Fields that may need additional analysis or extraction
            'may_need_analysis': [
                'PML %', 'CAT Exposure', 'Claims Experience (3 years)',
                'Share offered %', 'Surveyor\'s Report', 'Climate Change Risk',
                'ESG Risk Assessment', 'Premium Rates (%)', 'Reinsurance Deductions'
            ],
            
            # All 23 critical fields for validation
            'critical_fields': [
                'Reference Number', 'Date Received', 'Insured', 'Cedant/Reinsured', 'Broker',
                'Perils Covered', 'Geographical Limit', 'Situation of Risk/Voyage', 
                'Occupation of Insured', 'Main Activities', 'Total Sums Insured', 'Currency',
                'Excess/Retention', 'Premium Rates (%)', 'Period of Insurance', 'PML %',
                'CAT Exposure', 'Reinsurance Deductions', 'Claims Experience (3 years)',
                'Share offered %', 'Surveyor\'s Report', 'Climate Change Risk', 'ESG Risk Assessment'
            ]
        }    
   
    def extract_from_email(self, email_content: EmailContent) -> ExtractedData:
        """
        Extract structured data from email content and attachments
        
        Args:
            email_content: Parsed email with attachments
            
        Returns:
            ExtractedData with extracted information
        """
        extracted_data = ExtractedData()
        
        # Extract from email subject and body
        email_text = f"{email_content.subject} {email_content.body}"
        
        # Store attachment list for surveyor's report detection
        attachment_names = [att.filename for att in email_content.attachments if att.filename]
        extracted_data._attachment_list = attachment_names
        
        self._extract_from_text(email_text, extracted_data)
        
        # Set document type based on email classification
        if email_content.document_type:
            extracted_data.document_type = email_content.document_type
        
        # Extract from attachments
        for attachment in email_content.attachments:
            if attachment.processed_content:
                self._extract_from_attachment(attachment, extracted_data)
        
        # Extract parties from email metadata
        if email_content.sender:
            party = PartyData(
                name=email_content.sender,
                email=email_content.sender,
                role="sender"
            )
            extracted_data.parties.append(party)
        
        # Calculate overall confidence score
        extracted_data.confidence_score = self._calculate_confidence(extracted_data)
        
        return extracted_data
    
    def extract_from_email_with_ri_slip_prioritization(self, email_content: EmailContent) -> ExtractedData:
        """
        Extract structured data from email content with RI Slip identification and prioritization
        
        This method implements the RI Slip identification and prioritization logic:
        - Identifies RI Slips in PDF and DOCX files (primary formats)
        - Supports Excel, PowerPoint, and image formats as secondary options
        - Processes RI Slips first, then supporting documents
        - Validates RI Slip quality before processing
        
        Args:
            email_content: Parsed email with attachments
            
        Returns:
            ExtractedData with extracted information from prioritized documents
        """
        extracted_data = ExtractedData()
        
        # Extract from email subject and body first
        email_text = f"{email_content.subject} {email_content.body}"
        
        # Store attachment list for surveyor's report detection
        attachment_names = [att.filename for att in email_content.attachments if att.filename]
        extracted_data._attachment_list = attachment_names
        
        self._extract_from_text(email_text, extracted_data)
        
        # Create processing plan using RI Slip identification
        processing_plan = ri_slip_identifier.classify_email_attachments(email_content)
        
        # Add processing notes about the plan
        extracted_data.processing_notes.append(
            f"Email processing plan: {len(processing_plan.ri_slips)} RI Slips, "
            f"{len(processing_plan.supporting_documents)} supporting docs, "
            f"{len(processing_plan.low_priority_documents)} other docs"
        )
        
        # Process documents in priority order
        self._process_attachments_by_priority(processing_plan, email_content.attachments, extracted_data)
        
        # Extract parties from email metadata
        if email_content.sender:
            party = PartyData(
                name=email_content.sender,
                email=email_content.sender,
                role="sender"
            )
            extracted_data.parties.append(party)
        
        # Set document type based on RI Slip classification
        if processing_plan.ri_slips:
            primary_ri_slip = processing_plan.ri_slips[0]
            if primary_ri_slip.classification_result.ri_slip_type:
                extracted_data.document_type = primary_ri_slip.classification_result.ri_slip_type.value
        elif email_content.document_type:
            extracted_data.document_type = email_content.document_type
        
        # Calculate overall confidence score
        extracted_data.confidence_score = self._calculate_confidence(extracted_data)
        
        # Migrate legacy data to Analysis document format
        extracted_data.migrate_to_analysis_format()
        
        return extracted_data
    
    def _process_attachments_by_priority(self, processing_plan: EmailProcessingPlan, 
                                       attachments: List[AttachmentData], 
                                       extracted_data: ExtractedData):
        """
        Process attachments according to the prioritized processing plan
        
        Args:
            processing_plan: Email processing plan with prioritized documents
            attachments: List of email attachments
            extracted_data: ExtractedData object to populate
        """
        # Create filename to attachment mapping
        attachment_map = {att.filename: att for att in attachments}
        
        # Process RI Slips first (highest priority)
        for doc_classification in processing_plan.ri_slips:
            if not doc_classification.should_process:
                extracted_data.processing_notes.append(
                    f"Skipped RI Slip {doc_classification.attachment_filename}: "
                    f"{doc_classification.skip_reason}"
                )
                continue
            
            # Find the corresponding attachment
            attachment = attachment_map.get(doc_classification.attachment_filename)
            
            if attachment and attachment.processed_content:
                # Add note about RI Slip processing
                extracted_data.processing_notes.append(
                    f"Processing RI Slip: {attachment.filename} "
                    f"(Type: {doc_classification.classification_result.ri_slip_type}, "
                    f"Quality: {doc_classification.classification_result.quality_score:.2f})"
                )
                
                # Extract data with higher weight for RI Slips
                self._extract_from_ri_slip_attachment(attachment, extracted_data, 
                                                   doc_classification.classification_result)
        
        # Process supporting documents second
        for doc_classification in processing_plan.supporting_documents:
            if not doc_classification.should_process:
                continue
            
            attachment = attachment_map.get(doc_classification.attachment_filename)
            
            if attachment and attachment.processed_content:
                extracted_data.processing_notes.append(
                    f"Processing supporting document: {attachment.filename}"
                )
                self._extract_from_attachment(attachment, extracted_data)
        
        # Process low priority documents last (minimal extraction)
        for doc_classification in processing_plan.low_priority_documents:
            if not doc_classification.should_process:
                continue
            
            attachment = attachment_map.get(doc_classification.attachment_filename)
            
            if attachment and attachment.processed_content:
                # Only extract basic information from low priority documents
                self._extract_basic_info_from_attachment(attachment, extracted_data)
    
    def _extract_from_ri_slip_attachment(self, attachment: AttachmentData, 
                                       extracted_data: ExtractedData,
                                       identification_result: RISlipIdentificationResult):
        """
        Extract data from RI Slip attachment with enhanced processing
        
        Args:
            attachment: Attachment data with processed content
            extracted_data: ExtractedData object to populate
            identification_result: RI Slip identification result
        """
        if not attachment.processed_content:
            return
        
        # Standard extraction
        self._extract_from_attachment(attachment, extracted_data)
        
        # Enhanced extraction using RI Slip indicators
        indicators = identification_result.extracted_indicators
        
        # Use extracted indicators to enhance data quality
        if indicators.get('reference_numbers'):
            if not extracted_data.reference_number:
                extracted_data.reference_number = indicators['reference_numbers'][0]
                extracted_data.analysis_data.reference_number = indicators['reference_numbers'][0]
        
        if indicators.get('cedants'):
            if not extracted_data.cedant:
                extracted_data.cedant = indicators['cedants'][0]
                extracted_data.analysis_data.cedant_reinsured = indicators['cedants'][0]
        
        if indicators.get('sum_insured_amounts'):
            # Parse sum insured amounts
            for amount_str in indicators['sum_insured_amounts']:
                parsed_amount = self._parse_financial_amount(amount_str)
                if parsed_amount and not extracted_data.financial.sum_insured:
                    extracted_data.financial.sum_insured = parsed_amount['amount']
                    extracted_data.analysis_data.total_sums_insured = parsed_amount['amount']
                    if parsed_amount.get('currency'):
                        extracted_data.financial.currency = parsed_amount['currency']
                        extracted_data.analysis_data.currency = parsed_amount['currency']
        
        if indicators.get('percentages'):
            # Parse percentage values
            for percentage_str in indicators['percentages']:
                parsed_percentage = self._parse_percentage(percentage_str)
                if parsed_percentage and not extracted_data.analysis_data.share_offered_percentage:
                    extracted_data.analysis_data.share_offered_percentage = parsed_percentage
        
        # Set RI Slip specific metadata
        extracted_data.analysis_data.processing_notes.append(
            f"RI Slip identified: {identification_result.ri_slip_type} "
            f"(Confidence: {identification_result.confidence_score:.2f})"
        )
        
        # Update confidence based on RI Slip quality
        if identification_result.quality_score > 0.7:
            extracted_data.confidence_score = max(extracted_data.confidence_score, 0.8)
        elif identification_result.quality_score > 0.5:
            extracted_data.confidence_score = max(extracted_data.confidence_score, 0.6)
    
    def _extract_basic_info_from_attachment(self, attachment: AttachmentData, 
                                          extracted_data: ExtractedData):
        """
        Extract only basic information from low priority attachments
        
        Args:
            attachment: Attachment data with processed content
            extracted_data: ExtractedData object to populate
        """
        if not attachment.processed_content:
            return
        
        content = attachment.processed_content
        
        # Extract only basic text for reference
        if isinstance(content, OCRResult):
            # Only extract if text is short (likely contains key info)
            if len(content.text) < 1000:
                self._extract_from_text(content.text, extracted_data)
        elif isinstance(content, WordDocumentData):
            # Only extract from first few paragraphs
            if content.paragraphs:
                sample_text = ' '.join(content.paragraphs[:3])
                self._extract_from_text(sample_text, extracted_data)
        
        # Add note about minimal processing
        extracted_data.processing_notes.append(
            f"Minimal extraction from low priority document: {attachment.filename}"
        )
    
    def _extract_from_attachment(self, attachment: AttachmentData, extracted_data: ExtractedData):
        """Extract data from a single attachment"""
        if not attachment.processed_content:
            return
        
        content = attachment.processed_content
        
        if isinstance(content, OCRResult):
            # Extract from OCR text
            self._extract_from_text(content.text, extracted_data)
            
            # Use OCR regions for more precise extraction
            for region in content.regions:
                self._extract_from_text(region.text, extracted_data)
        
        elif isinstance(content, WordDocumentData):
            # Extract from Word document
            self._extract_from_text(content.text, extracted_data)
            
            # Extract from tables if present
            for table in content.tables:
                self._extract_from_table(table, extracted_data)
        
        elif isinstance(content, ExcelData):
            # Extract from Excel sheets
            for sheet_name, sheet_data in content.sheets.items():
                self._extract_from_excel_sheet(sheet_data, extracted_data)
        
        elif isinstance(content, PowerPointData):
            # Extract from PowerPoint slides
            self._extract_from_text(content.text, extracted_data)
    
    def _extract_from_text(self, text: str, extracted_data: ExtractedData):
        """Extract structured data from text using NLP and patterns"""
        if not text or not text.strip():
            return
        
        # Use Hugging Face models if available
        if self.models:
            self._extract_with_models(text, extracted_data)
        
        # Always use pattern-based extraction as fallback/supplement
        self._extract_with_patterns(text, extracted_data)
    
    def _extract_with_models(self, text: str, extracted_data: ExtractedData):
        """Extract data using Hugging Face models"""
        try:
            # Named Entity Recognition
            if 'ner' in self.models:
                entities = self.models['ner'](text)
                for entity in entities:
                    self._process_entity(entity, extracted_data)
            
            # Financial entity extraction
            if 'finbert' in self.models:
                fin_entities = self.models['finbert'](text)
                for entity in fin_entities:
                    self._process_financial_entity(entity, extracted_data)
            
            # Document classification
            if 'classifier' in self.models and not extracted_data.document_type:
                result = self.models['classifier'](text, self.keywords['document_types'])
                if result['scores'][0] > 0.5:
                    extracted_data.document_type = result['labels'][0]
            
        except Exception as e:
            logger.error(f"Error in model-based extraction: {str(e)}")
            extracted_data.processing_notes.append(f"Model extraction error: {str(e)}")
    
    def _extract_with_patterns(self, text: str, extracted_data: ExtractedData):
        """Extract data using regex patterns"""
        try:
            # Extract currency amounts
            amounts = self.patterns['currency_amount'].findall(text)
            for amount in amounts:
                self._process_amount(amount, extracted_data)
            
            # Extract policy numbers
            policy_matches = self.patterns['policy_number'].findall(text)
            if policy_matches and not extracted_data.reference_number:
                extracted_data.reference_number = policy_matches[0]
            
            # Extract dates
            date_matches = self.patterns['date'].findall(text)
            if date_matches and not extracted_data.date:
                extracted_data.date = self._parse_date(date_matches[0])
            
            # Extract emails
            emails = self.patterns['email'].findall(text)
            for email in emails:
                party = PartyData(email=email, role="contact")
                extracted_data.parties.append(party)
            
            # Extract locations
            locations = self.patterns['location'].findall(text)
            if locations and not extracted_data.risk.location:
                extracted_data.risk.location = locations[0]
            
            # Extract coverage types using keywords
            text_lower = text.lower()
            for coverage_type in self.keywords['coverage_types']:
                if coverage_type in text_lower:
                    extracted_data.financial.coverage_type = coverage_type
                    break
            
            # Extract line of business
            for lob in self.keywords['coverage_types']:
                if lob in text_lower:
                    extracted_data.line_of_business = lob
                    break
            
            # Extract treaty types
            for treaty_type in self.keywords['treaty_types']:
                if treaty_type in text_lower:
                    extracted_data.treaty_type = treaty_type
                    break
            
            # Extract parties using reinsurance terminology
            self._extract_parties_from_text(text, extracted_data)
            
            # NEW: Extract specialized reinsurance fields (Task 5.3)
            self._extract_specialized_reinsurance_fields(text, extracted_data)
            
        except Exception as e:
            logger.error(f"Error in pattern-based extraction: {str(e)}")
            extracted_data.processing_notes.append(f"Pattern extraction error: {str(e)}")
    
    def _extract_parties_from_text(self, text: str, extracted_data: ExtractedData):
        """Extract party information using reinsurance-specific terminology"""
        text_lower = text.lower()
        
        # Look for party information with context
        for party_type, keywords in self.keywords['party_keywords'].items():
            for keyword in keywords:
                # Pattern to find "keyword: company name" or "keyword company name"
                pattern = rf'{keyword}\s*:?\s*([A-Z][A-Za-z\s&.,()-]+(?:Ltd|Inc|Corp|Company|Insurance|Reinsurance|Group|plc|LLC|AG|SA|SE|NV|BV)?)'
                matches = re.findall(pattern, text, re.IGNORECASE)
                
                if matches:
                    company_name = matches[0].strip()
                    
                    # Clean up the company name
                    company_name = re.sub(r'\s+', ' ', company_name)
                    
                    # Assign to appropriate field
                    if party_type == 'reinsured' and not extracted_data.reinsured:
                        extracted_data.reinsured = company_name
                        extracted_data.cedant = company_name  # Same entity
                        extracted_data.insurer = company_name  # Same entity
                    elif party_type == 'reinsurer' and not extracted_data.reinsurer:
                        extracted_data.reinsurer = company_name
                    elif party_type == 'broker' and not extracted_data.broker:
                        extracted_data.broker = company_name
                    elif party_type == 'insured' and not extracted_data.insured:
                        extracted_data.insured = company_name
                    elif party_type == 'underwriter' and not extracted_data.underwriter:
                        extracted_data.underwriter = company_name
                    elif party_type == 'account_handler' and not extracted_data.account_handler:
                        extracted_data.account_handler = company_name
                    
                    # Also add to parties list
                    party = PartyData(
                        name=company_name,
                        role=party_type
                    )
                    extracted_data.parties.append(party)
    
    def _extract_from_table(self, table: List[List[str]], extracted_data: ExtractedData):
        """Extract data from table structure"""
        if not table or len(table) < 2:
            return
        
        # Convert table to DataFrame for easier processing
        try:
            df = pd.DataFrame(table[1:], columns=table[0])
            
            # Look for common reinsurance table structures
            for col in df.columns:
                col_lower = col.lower()
                
                if 'sum insured' in col_lower or 'limit' in col_lower:
                    values = df[col].dropna()
                    if not values.empty:
                        amount = self._parse_amount(str(values.iloc[0]))
                        if amount:
                            extracted_data.financial.sum_insured = amount
                
                elif 'premium' in col_lower:
                    values = df[col].dropna()
                    if not values.empty:
                        amount = self._parse_amount(str(values.iloc[0]))
                        if amount:
                            extracted_data.financial.premium = amount
                
                elif 'deductible' in col_lower or 'excess' in col_lower:
                    values = df[col].dropna()
                    if not values.empty:
                        amount = self._parse_amount(str(values.iloc[0]))
                        if amount:
                            extracted_data.financial.deductible = amount
                
                elif 'insured' in col_lower:
                    values = df[col].dropna()
                    if not values.empty and not extracted_data.insured:
                        extracted_data.insured = str(values.iloc[0])
                
                elif 'broker' in col_lower:
                    values = df[col].dropna()
                    if not values.empty and not extracted_data.broker:
                        extracted_data.broker = str(values.iloc[0])
        
        except Exception as e:
            logger.error(f"Error processing table: {str(e)}")
            extracted_data.processing_notes.append(f"Table processing error: {str(e)}")
    
    def _extract_from_excel_sheet(self, sheet_data: List[Dict[str, Any]], extracted_data: ExtractedData):
        """Extract data from Excel sheet data"""
        if not sheet_data:
            return
        
        try:
            df = pd.DataFrame(sheet_data)
            
            # Process each row and column for relevant data
            for col in df.columns:
                col_lower = str(col).lower()
                
                # Look for financial data
                if any(term in col_lower for term in ['sum insured', 'limit', 'coverage']):
                    values = df[col].dropna()
                    if not values.empty:
                        amount = self._parse_amount(str(values.iloc[0]))
                        if amount:
                            extracted_data.financial.sum_insured = amount
                
                # Look for party information
                elif 'insured' in col_lower:
                    values = df[col].dropna()
                    if not values.empty and not extracted_data.insured:
                        extracted_data.insured = str(values.iloc[0])
        
        except Exception as e:
            logger.error(f"Error processing Excel sheet: {str(e)}")
            extracted_data.processing_notes.append(f"Excel processing error: {str(e)}")
    
    def _process_entity(self, entity: Dict, extracted_data: ExtractedData):
        """Process named entity from NER model"""
        entity_type = entity.get('entity_group', '').upper()
        entity_text = entity.get('word', '').strip()
        
        if entity_type == 'PER' and not extracted_data.insured:
            # Person name - could be insured or contact
            extracted_data.insured = entity_text
        
        elif entity_type == 'ORG':
            # Organization - could be insured, broker, or reinsurer
            if not extracted_data.insured:
                extracted_data.insured = entity_text
            elif not extracted_data.broker:
                extracted_data.broker = entity_text
        
        elif entity_type == 'LOC':
            # Location
            if not extracted_data.risk.location:
                extracted_data.risk.location = entity_text
    
    def _process_financial_entity(self, entity: Dict, extracted_data: ExtractedData):
        """Process financial entity from FinBERT model"""
        entity_text = entity.get('word', '').strip()
        
        # Try to extract monetary amounts
        amount = self._parse_amount(entity_text)
        if amount:
            # Determine what type of amount this might be based on context
            if not extracted_data.financial.sum_insured:
                extracted_data.financial.sum_insured = amount
            elif not extracted_data.financial.premium:
                extracted_data.financial.premium = amount
    
    def _process_amount(self, amount_text: str, extracted_data: ExtractedData):
        """Process monetary amount text"""
        amount = self._parse_amount(amount_text)
        if amount:
            # Extract currency
            currency_match = re.search(r'(USD|EUR|GBP|CAD|AUD|JPY|CHF)', amount_text, re.IGNORECASE)
            if currency_match:
                extracted_data.financial.currency = currency_match.group(1).upper()
            
            # Assign to appropriate field based on context
            if not extracted_data.financial.sum_insured:
                extracted_data.financial.sum_insured = amount
            elif not extracted_data.financial.premium:
                extracted_data.financial.premium = amount
            elif not extracted_data.financial.limit:
                extracted_data.financial.limit = amount
    
    def _parse_amount(self, amount_text: str) -> Optional[Decimal]:
        """Parse monetary amount from text"""
        if not amount_text:
            return None
        
        try:
            # Remove currency symbols and clean up
            cleaned = re.sub(r'[^\d.,]', '', amount_text)
            cleaned = cleaned.replace(',', '')
            
            if cleaned:
                return Decimal(cleaned)
        except Exception:
            pass
        
        return None
    
    def _parse_date(self, date_text: str) -> Optional[datetime]:
        """Parse date from text"""
        if not date_text:
            return None
        
        # Try different date formats
        formats = [
            '%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d',
            '%m-%d-%Y', '%d-%m-%Y', '%Y/%m/%d',
            '%d %b %Y', '%d %B %Y', '%b %d, %Y'
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(date_text, fmt)
            except ValueError:
                continue
        
        return None
    
    def _parse_financial_amount(self, amount_text: str) -> Optional[Dict[str, Any]]:
        """
        Parse financial amount with currency from text
        
        Args:
            amount_text: Text containing financial amount
            
        Returns:
            Dictionary with 'amount' and 'currency' keys, or None
        """
        if not amount_text:
            return None
        
        # Pattern to match currency and amount
        pattern = r'([A-Z]{3})\s*[\$€£¥]?\s*([\d,]+(?:\.\d{2})?)'
        match = re.search(pattern, amount_text, re.IGNORECASE)
        
        if match:
            currency = match.group(1).upper()
            amount_str = match.group(2).replace(',', '')
            
            try:
                amount = Decimal(amount_str)
                return {'amount': amount, 'currency': currency}
            except (ValueError, TypeError):
                pass
        
        # Try without currency
        amount_pattern = r'[\d,]+(?:\.\d{2})?'
        amount_match = re.search(amount_pattern, amount_text)
        
        if amount_match:
            amount_str = amount_match.group().replace(',', '')
            try:
                amount = Decimal(amount_str)
                return {'amount': amount, 'currency': None}
            except (ValueError, TypeError):
                pass
        
        return None
    
    def _parse_percentage(self, percentage_text: str) -> Optional[Decimal]:
        """
        Parse percentage value from text
        
        Args:
            percentage_text: Text containing percentage
            
        Returns:
            Decimal percentage value, or None
        """
        if not percentage_text:
            return None
        
        # Remove % symbol and extract number
        clean_text = percentage_text.replace('%', '').strip()
        
        # Pattern to match decimal numbers
        pattern = r'(\d+(?:\.\d+)?)'
        match = re.search(pattern, clean_text)
        
        if match:
            try:
                percentage = Decimal(match.group(1))
                # Ensure percentage is within valid range
                if 0 <= percentage <= 100:
                    return percentage
            except (ValueError, TypeError):
                pass
        
        return None
    
    def _extract_specialized_reinsurance_fields(self, text: str, extracted_data: ExtractedData):
        """
        Extract specialized reinsurance fields using enhanced patterns (Task 5.3)
        
        This method implements extraction for:
        - Basic fields: Insured, Cedant, Broker, Perils Covered
        - Geographic and risk patterns: Geographical Limit, Situation of Risk/Voyage
        - Business patterns: Occupation of Insured, Main Activities
        - Financial patterns: Total Sums Insured, Excess, Retention, Premium Rates
        
        Args:
            text: Text to extract from
            extracted_data: ExtractedData object to populate
        """
        try:
            # 1. BASIC FIELDS EXTRACTION
            
            # Extract Insured (end policyholder)
            if not extracted_data.analysis_data.insured_name:
                insured_matches = self.patterns['insured_pattern'].findall(text)
                if insured_matches:
                    insured_name = self._clean_company_name(insured_matches[0])
                    extracted_data.analysis_data.insured_name = insured_name
                    extracted_data.insured = insured_name  # Legacy field
                    extracted_data.processing_notes.append(f"Extracted insured: {insured_name}")
            
            # Extract Cedant/Reinsured (ceding company)
            if not extracted_data.analysis_data.cedant_reinsured:
                cedant_matches = self.patterns['cedant_pattern'].findall(text)
                if cedant_matches:
                    cedant_name = self._clean_company_name(cedant_matches[0])
                    extracted_data.analysis_data.cedant_reinsured = cedant_name
                    extracted_data.cedant = cedant_name  # Legacy field
                    extracted_data.reinsured = cedant_name  # Legacy field
                    extracted_data.processing_notes.append(f"Extracted cedant: {cedant_name}")
            
            # Extract Broker
            if not extracted_data.analysis_data.broker_name:
                broker_matches = self.patterns['broker_pattern'].findall(text)
                if broker_matches:
                    broker_name = self._clean_company_name(broker_matches[0])
                    extracted_data.analysis_data.broker_name = broker_name
                    extracted_data.broker = broker_name  # Legacy field
                    extracted_data.processing_notes.append(f"Extracted broker: {broker_name}")
            
            # Extract Perils Covered
            if not extracted_data.analysis_data.perils_covered:
                perils_matches = self.patterns['perils_pattern'].findall(text)
                if perils_matches:
                    perils_text = self._clean_text_field(perils_matches[0])
                    extracted_data.analysis_data.perils_covered = perils_text
                    extracted_data.processing_notes.append(f"Extracted perils: {perils_text[:50]}...")
            
            # 2. GEOGRAPHIC AND RISK PATTERNS
            
            # Extract Geographical Limit
            if not extracted_data.analysis_data.geographical_limit:
                geo_limit_matches = self.patterns['geographical_limit_pattern'].findall(text)
                if geo_limit_matches:
                    geo_limit = self._clean_text_field(geo_limit_matches[0])
                    extracted_data.analysis_data.geographical_limit = geo_limit
                    extracted_data.processing_notes.append(f"Extracted geographical limit: {geo_limit[:50]}...")
            
            # Extract Situation of Risk/Voyage
            if not extracted_data.analysis_data.situation_of_risk:
                situation_matches = self.patterns['situation_of_risk_pattern'].findall(text)
                if situation_matches:
                    situation = self._clean_text_field(situation_matches[0])
                    extracted_data.analysis_data.situation_of_risk = situation
                    extracted_data.risk.location = situation  # Legacy field
                    extracted_data.processing_notes.append(f"Extracted situation of risk: {situation[:50]}...")
            
            # 3. BUSINESS PATTERNS
            
            # Extract Occupation of Insured
            if not extracted_data.analysis_data.occupation_of_insured:
                occupation_matches = self.patterns['occupation_pattern'].findall(text)
                if occupation_matches:
                    occupation = self._clean_text_field(occupation_matches[0])
                    extracted_data.analysis_data.occupation_of_insured = occupation
                    extracted_data.risk.industry = occupation  # Legacy field
                    extracted_data.processing_notes.append(f"Extracted occupation: {occupation[:50]}...")
            
            # Extract Main Activities
            if not extracted_data.analysis_data.main_activities:
                activities_matches = self.patterns['main_activities_pattern'].findall(text)
                if activities_matches:
                    activities = self._clean_text_field(activities_matches[0])
                    extracted_data.analysis_data.main_activities = activities
                    extracted_data.processing_notes.append(f"Extracted main activities: {activities[:50]}...")
            
            # 4. FINANCIAL PATTERNS
            
            # Extract Total Sums Insured
            if not extracted_data.analysis_data.total_sums_insured:
                tsi_matches = self.patterns['total_sum_insured_pattern'].findall(text)
                if tsi_matches:
                    currency, amount_str = tsi_matches[0]
                    parsed_amount = self._parse_financial_amount_with_multiplier(amount_str, currency)
                    if parsed_amount:
                        extracted_data.analysis_data.total_sums_insured = parsed_amount['amount']
                        if parsed_amount.get('currency'):
                            extracted_data.analysis_data.currency = parsed_amount['currency']
                            extracted_data.financial.currency = parsed_amount['currency']  # Legacy
                        extracted_data.financial.sum_insured = parsed_amount['amount']  # Legacy
                        extracted_data.processing_notes.append(f"Extracted TSI: {parsed_amount['currency']} {parsed_amount['amount']}")
            
            # Extract Excess/Retention
            if not extracted_data.analysis_data.excess_retention:
                excess_matches = self.patterns['excess_retention_pattern'].findall(text)
                if excess_matches:
                    currency, amount_str = excess_matches[0]
                    parsed_amount = self._parse_financial_amount_with_multiplier(amount_str, currency)
                    if parsed_amount:
                        extracted_data.analysis_data.excess_retention = parsed_amount['amount']
                        extracted_data.financial.deductible = parsed_amount['amount']  # Legacy
                        extracted_data.processing_notes.append(f"Extracted excess/retention: {parsed_amount.get('currency', '')} {parsed_amount['amount']}")
            
            # Extract Premium Rates
            if not extracted_data.analysis_data.premium_rates:
                premium_matches = self.patterns['premium_rates_pattern'].findall(text)
                if premium_matches:
                    rate_str = premium_matches[0]
                    try:
                        rate = Decimal(rate_str)
                        # Ensure rate is reasonable (0-100%)
                        if 0 <= rate <= 100:
                            extracted_data.analysis_data.premium_rates = rate
                            extracted_data.processing_notes.append(f"Extracted premium rate: {rate}%")
                    except (ValueError, TypeError):
                        pass
            
            # Extract Period of Insurance
            if not extracted_data.analysis_data.period_of_insurance:
                period_matches = self.patterns['period_of_insurance_pattern'].findall(text)
                if period_matches:
                    period = self._clean_text_field(period_matches[0])
                    extracted_data.analysis_data.period_of_insurance = period
                    extracted_data.financial.policy_period = period  # Legacy
                    extracted_data.processing_notes.append(f"Extracted period: {period}")
            
            # 5. ENHANCED KEYWORD-BASED EXTRACTION
            
            # Extract additional perils using keyword matching
            if not extracted_data.analysis_data.perils_covered:
                detected_perils = []
                text_lower = text.lower()
                for peril in self.keywords['perils_keywords']:
                    if peril in text_lower:
                        detected_perils.append(peril)
                
                if detected_perils:
                    perils_text = ', '.join(detected_perils[:10])  # Limit to first 10 found
                    extracted_data.analysis_data.perils_covered = perils_text
                    extracted_data.processing_notes.append(f"Extracted perils from keywords: {perils_text}")
            
            # Extract occupation using keyword matching if pattern didn't work
            if not extracted_data.analysis_data.occupation_of_insured:
                text_lower = text.lower()
                for occupation in self.keywords['occupation_keywords']:
                    if occupation in text_lower:
                        extracted_data.analysis_data.occupation_of_insured = occupation.title()
                        extracted_data.risk.industry = occupation.title()  # Legacy
                        extracted_data.processing_notes.append(f"Extracted occupation from keywords: {occupation}")
                        break
            
            # 6. ADVANCED REINSURANCE-SPECIFIC FIELDS (Task 5.4)
            
            # Extract PML % with validation
            if not extracted_data.analysis_data.pml_percentage:
                pml_value, pml_confidence = self.extract_pml_percentage(text)
                if pml_value:
                    extracted_data.analysis_data.pml_percentage = pml_value
                    extracted_data.analysis_data.field_confidence_scores['pml_percentage'] = pml_confidence
                    extracted_data.processing_notes.append(f"Extracted PML %: {pml_value}% (confidence: {pml_confidence:.2f})")
            
            # Extract CAT Exposure with risk validation
            if not extracted_data.analysis_data.cat_exposure:
                cat_exposure, cat_confidence = self.extract_cat_exposure(text)
                if cat_exposure:
                    extracted_data.analysis_data.cat_exposure = cat_exposure
                    extracted_data.analysis_data.field_confidence_scores['cat_exposure'] = cat_confidence
                    extracted_data.processing_notes.append(f"Extracted CAT exposure: {cat_exposure[:50]}... (confidence: {cat_confidence:.2f})")
            
            # Extract Period of Insurance (enhanced)
            if not extracted_data.analysis_data.period_of_insurance:
                period, period_confidence = self.extract_period_of_insurance(text)
                if period:
                    extracted_data.analysis_data.period_of_insurance = period
                    extracted_data.analysis_data.field_confidence_scores['period_of_insurance'] = period_confidence
                    extracted_data.financial.policy_period = period  # Legacy
                    extracted_data.processing_notes.append(f"Extracted period of insurance: {period} (confidence: {period_confidence:.2f})")
            
            # Extract Reinsurance Deductions
            if not extracted_data.analysis_data.reinsurance_deductions:
                deductions, deductions_confidence = self.extract_reinsurance_deductions(text)
                if deductions:
                    extracted_data.analysis_data.reinsurance_deductions = deductions
                    extracted_data.analysis_data.field_confidence_scores['reinsurance_deductions'] = deductions_confidence
                    extracted_data.processing_notes.append(f"Extracted reinsurance deductions: {deductions} (confidence: {deductions_confidence:.2f})")
            
            # Extract Claims Experience (3 years)
            if not extracted_data.analysis_data.claims_experience_3_years:
                claims_exp, claims_confidence = self.extract_claims_experience_3_years(text)
                if claims_exp:
                    extracted_data.analysis_data.claims_experience_3_years = claims_exp
                    extracted_data.analysis_data.field_confidence_scores['claims_experience_3_years'] = claims_confidence
                    extracted_data.processing_notes.append(f"Extracted claims experience: {claims_exp[:50]}... (confidence: {claims_confidence:.2f})")
            
            # Extract Share offered Percentage
            if not extracted_data.analysis_data.share_offered_percentage:
                share_pct, share_confidence = self.extract_share_offered_percentage(text)
                if share_pct:
                    extracted_data.analysis_data.share_offered_percentage = share_pct
                    extracted_data.analysis_data.field_confidence_scores['share_offered_percentage'] = share_confidence
                    extracted_data.processing_notes.append(f"Extracted share offered %: {share_pct}% (confidence: {share_confidence:.2f})")
            
            # Extract Surveyor's Report (with attachment detection)
            if not extracted_data.analysis_data.surveyors_report:
                # Get attachment list if available (from email processing)
                attachments = getattr(extracted_data, '_attachment_list', [])
                survey_report, survey_confidence = self.extract_surveyors_report(text, attachments)
                if survey_report:
                    extracted_data.analysis_data.surveyors_report = survey_report
                    extracted_data.analysis_data.field_confidence_scores['surveyors_report'] = survey_confidence
                    extracted_data.processing_notes.append(f"Extracted surveyor's report: {survey_report[:50]}... (confidence: {survey_confidence:.2f})")
            
            # Extract Climate Change Risk Assessment
            if not extracted_data.analysis_data.climate_change_risk:
                climate_risk, climate_confidence = self.extract_climate_change_risk(text)
                if climate_risk:
                    extracted_data.analysis_data.climate_change_risk = climate_risk
                    extracted_data.analysis_data.field_confidence_scores['climate_change_risk'] = climate_confidence
                    extracted_data.processing_notes.append(f"Extracted climate change risk: {climate_risk[:50]}... (confidence: {climate_confidence:.2f})")
            
            # Extract ESG Risk Assessment
            if not extracted_data.analysis_data.esg_risk_assessment:
                esg_risk, esg_confidence = self.extract_esg_risk_assessment(text)
                if esg_risk:
                    extracted_data.analysis_data.esg_risk_assessment = esg_risk
                    extracted_data.analysis_data.field_confidence_scores['esg_risk_assessment'] = esg_confidence
                    extracted_data.processing_notes.append(f"Extracted ESG risk assessment: {esg_risk[:50]}... (confidence: {esg_confidence:.2f})")
            
        except Exception as e:
            logger.error(f"Error in specialized reinsurance field extraction: {str(e)}")
            extracted_data.processing_notes.append(f"Specialized extraction error: {str(e)}")
    
    def _clean_company_name(self, company_name: str) -> str:
        """
        Clean and standardize company name
        
        Args:
            company_name: Raw company name text
            
        Returns:
            Cleaned company name
        """
        if not company_name:
            return ""
        
        # Remove extra whitespace and normalize
        cleaned = re.sub(r'\s+', ' ', company_name.strip())
        
        # Remove common prefixes/suffixes that might be extraction artifacts
        cleaned = re.sub(r'^(the\s+|a\s+)', '', cleaned, flags=re.IGNORECASE)
        cleaned = re.sub(r'\s+(and\s+associates|&\s+co\.?|&\s+company)$', '', cleaned, flags=re.IGNORECASE)
        
        # Capitalize properly
        cleaned = cleaned.title()
        
        # Fix common abbreviations
        cleaned = re.sub(r'\bLtd\b', 'Ltd.', cleaned)
        cleaned = re.sub(r'\bInc\b', 'Inc.', cleaned)
        cleaned = re.sub(r'\bCorp\b', 'Corp.', cleaned)
        cleaned = re.sub(r'\bLlc\b', 'LLC', cleaned)
        
        return cleaned
    
    def _clean_text_field(self, text: str) -> str:
        """
        Clean and standardize text field content
        
        Args:
            text: Raw text content
            
        Returns:
            Cleaned text
        """
        if not text:
            return ""
        
        # Remove extra whitespace and normalize
        cleaned = re.sub(r'\s+', ' ', text.strip())
        
        # Remove common extraction artifacts
        cleaned = re.sub(r'^[:\-\s]+', '', cleaned)
        cleaned = re.sub(r'[:\-\s]+$', '', cleaned)
        
        # Limit length to reasonable size
        if len(cleaned) > 500:
            cleaned = cleaned[:497] + "..."
        
        return cleaned
    
    def _parse_financial_amount_with_multiplier(self, amount_str: str, currency: str = None) -> Optional[Dict[str, Any]]:
        """
        Parse financial amount with multiplier support (million, thousand, etc.)
        
        Args:
            amount_str: Amount string (e.g., "1.5", "2,500")
            currency: Currency code (optional)
            
        Returns:
            Dictionary with 'amount' and 'currency' keys, or None
        """
        if not amount_str:
            return None
        
        try:
            # Clean the amount string
            clean_amount = amount_str.replace(',', '').strip()
            
            # Parse the base amount
            base_amount = Decimal(clean_amount)
            
            # Apply multipliers based on context (this would need to be enhanced
            # to detect multipliers from surrounding text)
            final_amount = base_amount
            
            return {
                'amount': final_amount,
                'currency': currency.upper() if currency else None
            }
            
        except (ValueError, TypeError):
            return None
    
    def _calculate_confidence(self, extracted_data: ExtractedData) -> float:
        """Calculate confidence score for extracted data"""
        score = 0.0
        total_fields = 0
        
        # Check key fields and assign weights (focusing on commonly available fields)
        fields_weights = {
            'reference_number': 0.15,
            'insured': 0.12,
            'financial.sum_insured': 0.18,
            'financial.currency': 0.08,
            'risk.location': 0.10,
            'document_type': 0.10,
            'date': 0.08,
            'broker': 0.09,
            'financial.coverage_type': 0.05,
            'line_of_business': 0.05
        }
        
        # Track data completeness
        extracted_data.data_completeness = {}
        
        for field, weight in fields_weights.items():
            total_fields += weight
            
            if '.' in field:
                # Nested field
                obj, attr = field.split('.')
                obj_value = getattr(extracted_data, obj, None)
                field_value = obj_value and getattr(obj_value, attr, None)
            else:
                # Direct field
                field_value = getattr(extracted_data, field, None)
            
            # Record completeness
            extracted_data.data_completeness[field] = bool(field_value)
            
            if field_value:
                score += weight
        
        # Add notes about missing key data
        missing_fields = [field for field, present in extracted_data.data_completeness.items() if not present]
        if missing_fields:
            extracted_data.processing_notes.append(f"Missing key fields: {', '.join(missing_fields)}")
        
        return min(score / total_fields if total_fields > 0 else 0.0, 1.0)
    
    # NEW: Advanced reinsurance-specific field extraction methods (Task 5.4)
    
    def extract_pml_percentage(self, text: str) -> Tuple[Optional[Decimal], float]:
        """
        Extract PML (Probable Maximum Loss) percentage with validation
        
        Args:
            text: Text to extract from
            
        Returns:
            Tuple of (PML percentage, confidence score)
        """
        confidence = 0.0
        pml_value = None
        
        try:
            # Try pattern matching first
            matches = self.patterns['pml_percentage_pattern'].findall(text)
            if matches:
                pml_str = matches[0].replace(',', '')
                pml_value = Decimal(pml_str)
                confidence = 0.9
                
                # Validate percentage range
                if pml_value < 0 or pml_value > 100:
                    confidence = 0.3  # Low confidence for out-of-range values
                    if pml_value > 100:
                        pml_value = min(pml_value, Decimal('100'))  # Cap at 100%
            
            # Try keyword-based extraction if pattern fails
            if not pml_value:
                for keyword in self.keywords['pml_keywords']:
                    if keyword.lower() in text.lower():
                        # Look for percentage near keyword
                        keyword_pos = text.lower().find(keyword.lower())
                        surrounding_text = text[max(0, keyword_pos-50):keyword_pos+100]
                        
                        percentage_matches = self.patterns['percentage'].findall(surrounding_text)
                        if percentage_matches:
                            pml_str = percentage_matches[0].replace('%', '').replace(',', '')
                            pml_value = Decimal(pml_str)
                            confidence = 0.7
                            break
            
            # Additional validation for reinsurance context
            if pml_value and confidence > 0.5:
                # Check if value is reasonable for reinsurance (typically 1-100%)
                if pml_value >= 1 and pml_value <= 100:
                    confidence = min(confidence + 0.1, 1.0)
                elif pml_value < 1:
                    confidence = max(confidence - 0.2, 0.1)
            
        except (ValueError, TypeError) as e:
            logger.warning(f"Error extracting PML percentage: {str(e)}")
            confidence = 0.0
        
        return pml_value, confidence
    
    def extract_cat_exposure(self, text: str) -> Tuple[Optional[str], float]:
        """
        Extract catastrophe exposure details with risk validation
        
        Args:
            text: Text to extract from
            
        Returns:
            Tuple of (CAT exposure description, confidence score)
        """
        confidence = 0.0
        cat_exposure = None
        
        try:
            # Try pattern matching first
            matches = self.patterns['cat_exposure_pattern'].findall(text)
            if matches:
                cat_exposure = matches[0].strip()
                confidence = 0.8
            
            # Try keyword-based extraction
            if not cat_exposure:
                cat_keywords_found = []
                for keyword in self.keywords['cat_exposure_keywords']:
                    if keyword.lower() in text.lower():
                        cat_keywords_found.append(keyword)
                
                if cat_keywords_found:
                    # Extract surrounding context
                    for keyword in cat_keywords_found[:3]:  # Limit to first 3 matches
                        keyword_pos = text.lower().find(keyword.lower())
                        surrounding_text = text[max(0, keyword_pos-30):keyword_pos+150]
                        
                        if not cat_exposure:
                            cat_exposure = surrounding_text.strip()
                            confidence = 0.6
                        else:
                            cat_exposure += f"; {surrounding_text.strip()}"
                            confidence = min(confidence + 0.1, 0.9)
            
            # Clean up extracted text
            if cat_exposure:
                # Remove excessive whitespace and clean up
                cat_exposure = ' '.join(cat_exposure.split())
                
                # Truncate if too long
                if len(cat_exposure) > 300:
                    cat_exposure = cat_exposure[:297] + "..."
                
                # Increase confidence if specific catastrophe models mentioned
                model_keywords = ['rms', 'air', 'eqecat', 'karen clark', 'model']
                if any(model.lower() in cat_exposure.lower() for model in model_keywords):
                    confidence = min(confidence + 0.15, 1.0)
        
        except Exception as e:
            logger.warning(f"Error extracting CAT exposure: {str(e)}")
            confidence = 0.0
        
        return cat_exposure, confidence
    
    def extract_period_of_insurance(self, text: str) -> Tuple[Optional[str], float]:
        """
        Extract period of insurance with enhanced date parsing
        
        Args:
            text: Text to extract from
            
        Returns:
            Tuple of (period description, confidence score)
        """
        confidence = 0.0
        period = None
        
        try:
            # Try detailed pattern matching first
            matches = self.patterns['period_of_insurance_detailed_pattern'].findall(text)
            if matches:
                if isinstance(matches[0], tuple) and len(matches[0]) >= 2:
                    start_date, end_date = matches[0][:2]
                    if start_date and end_date:
                        period = f"{start_date} to {end_date}"
                        confidence = 0.9
                    elif start_date:
                        period = f"From {start_date}"
                        confidence = 0.7
                else:
                    period = str(matches[0])
                    confidence = 0.8
            
            # Try original pattern if detailed fails
            if not period:
                matches = self.patterns['period_of_insurance_pattern'].findall(text)
                if matches:
                    period = matches[0]
                    confidence = 0.7
            
            # Try keyword-based extraction
            if not period:
                for keyword in self.keywords['financial_keywords']['period']:
                    if keyword.lower() in text.lower():
                        keyword_pos = text.lower().find(keyword.lower())
                        surrounding_text = text[max(0, keyword_pos-20):keyword_pos+80]
                        
                        # Look for dates in surrounding text
                        date_matches = self.patterns['date'].findall(surrounding_text)
                        if date_matches:
                            if len(date_matches) >= 2:
                                period = f"{date_matches[0]} to {date_matches[1]}"
                            else:
                                period = date_matches[0]
                            confidence = 0.6
                            break
            
            # Clean up and validate
            if period:
                period = ' '.join(period.split())
                
                # Check for common insurance periods
                common_periods = ['12 months', '1 year', '24 months', '2 years', '36 months', '3 years']
                if any(cp in period.lower() for cp in common_periods):
                    confidence = min(confidence + 0.1, 1.0)
        
        except Exception as e:
            logger.warning(f"Error extracting period of insurance: {str(e)}")
            confidence = 0.0
        
        return period, confidence
    
    def extract_reinsurance_deductions(self, text: str) -> Tuple[Optional[Decimal], float]:
        """
        Extract reinsurance deductions with financial validation
        
        Args:
            text: Text to extract from
            
        Returns:
            Tuple of (deduction amount, confidence score)
        """
        confidence = 0.0
        deductions = None
        
        try:
            # Try pattern matching first
            matches = self.patterns['reinsurance_deductions_pattern'].findall(text)
            if matches:
                for match in matches:
                    if isinstance(match, tuple):
                        # Extract currency and amount
                        currency, amount_str = match
                        if amount_str:
                            amount_str = amount_str.replace(',', '')
                            deductions = Decimal(amount_str)
                            confidence = 0.8
                            break
                    else:
                        # Direct percentage or amount
                        amount_str = str(match).replace(',', '').replace('%', '')
                        deductions = Decimal(amount_str)
                        confidence = 0.7
                        break
            
            # Try keyword-based extraction
            if not deductions:
                for keyword in self.keywords['reinsurance_deductions_keywords']:
                    if keyword.lower() in text.lower():
                        keyword_pos = text.lower().find(keyword.lower())
                        surrounding_text = text[max(0, keyword_pos-30):keyword_pos+100]
                        
                        # Look for amounts or percentages
                        amount_matches = self.patterns['currency_amount'].findall(surrounding_text)
                        if amount_matches:
                            amount_str = amount_matches[0].split()[-1].replace(',', '')
                            deductions = Decimal(amount_str)
                            confidence = 0.6
                            break
                        
                        percentage_matches = self.patterns['percentage'].findall(surrounding_text)
                        if percentage_matches:
                            pct_str = percentage_matches[0].replace('%', '').replace(',', '')
                            deductions = Decimal(pct_str)
                            confidence = 0.6
                            break
            
            # Validate amount
            if deductions:
                # Check if reasonable for reinsurance deductions (typically 0-50%)
                if deductions >= 0 and deductions <= 50:
                    confidence = min(confidence + 0.1, 1.0)
                elif deductions > 50 and deductions <= 100:
                    confidence = max(confidence - 0.1, 0.3)
                elif deductions > 100:
                    # Might be absolute amount rather than percentage
                    confidence = max(confidence - 0.2, 0.2)
        
        except (ValueError, TypeError) as e:
            logger.warning(f"Error extracting reinsurance deductions: {str(e)}")
            confidence = 0.0
        
        return deductions, confidence
    
    def extract_claims_experience_3_years(self, text: str) -> Tuple[Optional[str], float]:
        """
        Extract 3-year claims experience with validation
        
        Args:
            text: Text to extract from
            
        Returns:
            Tuple of (claims experience description, confidence score)
        """
        confidence = 0.0
        claims_experience = None
        
        try:
            # Try pattern matching first
            matches = self.patterns['claims_experience_pattern'].findall(text)
            if matches:
                claims_experience = matches[0].strip()
                confidence = 0.8
            
            # Try keyword-based extraction
            if not claims_experience:
                experience_keywords_found = []
                for keyword in self.keywords['claims_experience_keywords']:
                    if keyword.lower() in text.lower():
                        experience_keywords_found.append(keyword)
                
                if experience_keywords_found:
                    # Extract surrounding context
                    for keyword in experience_keywords_found[:2]:  # Limit to first 2 matches
                        keyword_pos = text.lower().find(keyword.lower())
                        surrounding_text = text[max(0, keyword_pos-50):keyword_pos+200]
                        
                        if not claims_experience:
                            claims_experience = surrounding_text.strip()
                            confidence = 0.6
                        else:
                            claims_experience += f"; {surrounding_text.strip()}"
                            confidence = min(confidence + 0.1, 0.9)
            
            # Clean up and validate
            if claims_experience:
                # Remove excessive whitespace
                claims_experience = ' '.join(claims_experience.split())
                
                # Truncate if too long
                if len(claims_experience) > 1000:
                    claims_experience = claims_experience[:997] + "..."
                
                # Increase confidence for specific indicators
                positive_indicators = ['nil', 'none', 'no claims', 'clean', 'excellent', 'good']
                negative_indicators = ['poor', 'bad', 'adverse', 'significant losses']
                
                claims_lower = claims_experience.lower()
                if any(indicator in claims_lower for indicator in positive_indicators):
                    confidence = min(confidence + 0.15, 1.0)
                elif any(indicator in claims_lower for indicator in negative_indicators):
                    confidence = min(confidence + 0.1, 0.9)
                
                # Check for specific amounts or years
                if any(year in claims_experience for year in ['2021', '2022', '2023', '2024']):
                    confidence = min(confidence + 0.1, 1.0)
        
        except Exception as e:
            logger.warning(f"Error extracting claims experience: {str(e)}")
            confidence = 0.0
        
        return claims_experience, confidence
    
    def extract_share_offered_percentage(self, text: str) -> Tuple[Optional[Decimal], float]:
        """
        Extract share offered percentage with validation
        
        Args:
            text: Text to extract from
            
        Returns:
            Tuple of (share percentage, confidence score)
        """
        confidence = 0.0
        share_percentage = None
        
        try:
            # Try pattern matching first
            matches = self.patterns['share_offered_percentage_pattern'].findall(text)
            if matches:
                share_str = matches[0].replace(',', '')
                share_percentage = Decimal(share_str)
                confidence = 0.9
                
                # Validate percentage range
                if share_percentage < 0 or share_percentage > 100:
                    confidence = 0.3
                    if share_percentage > 100:
                        share_percentage = min(share_percentage, Decimal('100'))
            
            # Try keyword-based extraction
            if not share_percentage:
                for keyword in self.keywords['share_offered_keywords']:
                    if keyword.lower() in text.lower():
                        keyword_pos = text.lower().find(keyword.lower())
                        surrounding_text = text[max(0, keyword_pos-30):keyword_pos+80]
                        
                        percentage_matches = self.patterns['percentage'].findall(surrounding_text)
                        if percentage_matches:
                            share_str = percentage_matches[0].replace('%', '').replace(',', '')
                            share_percentage = Decimal(share_str)
                            confidence = 0.7
                            break
            
            # Additional validation for reinsurance context
            if share_percentage and confidence > 0.5:
                # Check if value is reasonable for reinsurance share (typically 1-100%)
                if share_percentage >= 1 and share_percentage <= 100:
                    confidence = min(confidence + 0.1, 1.0)
                elif share_percentage < 1 and share_percentage > 0:
                    # Might be fractional share (e.g., 0.25 for 25%)
                    if share_percentage < 1:
                        share_percentage = share_percentage * 100
                        confidence = max(confidence - 0.1, 0.5)
        
        except (ValueError, TypeError) as e:
            logger.warning(f"Error extracting share offered percentage: {str(e)}")
            confidence = 0.0
        
        return share_percentage, confidence
    
    def extract_surveyors_report(self, text: str, attachments: List[str] = None) -> Tuple[Optional[str], float]:
        """
        Extract surveyor's report details and detect attachment links
        
        Args:
            text: Text to extract from
            attachments: List of attachment filenames (optional)
            
        Returns:
            Tuple of (surveyor's report description, confidence score)
        """
        confidence = 0.0
        surveyors_report = None
        
        try:
            # Try pattern matching first
            matches = self.patterns['surveyors_report_pattern'].findall(text)
            if matches:
                surveyors_report = matches[0].strip()
                confidence = 0.8
            
            # Check for attachment references
            attachment_matches = self.patterns['surveyors_report_attachment_pattern'].findall(text)
            if attachment_matches and attachments:
                # Cross-reference with actual attachments
                for attachment_ref in attachment_matches:
                    if any(attachment_ref.lower() in att.lower() for att in attachments):
                        if surveyors_report:
                            surveyors_report += f"; Attachment: {attachment_ref}"
                        else:
                            surveyors_report = f"Attachment: {attachment_ref}"
                        confidence = min(confidence + 0.2, 1.0)
            
            # Try keyword-based extraction
            if not surveyors_report:
                survey_keywords_found = []
                for keyword in self.keywords['surveyors_report_keywords']:
                    if keyword.lower() in text.lower():
                        survey_keywords_found.append(keyword)
                
                if survey_keywords_found:
                    # Extract surrounding context
                    for keyword in survey_keywords_found[:2]:
                        keyword_pos = text.lower().find(keyword.lower())
                        surrounding_text = text[max(0, keyword_pos-30):keyword_pos+150]
                        
                        if not surveyors_report:
                            surveyors_report = surrounding_text.strip()
                            confidence = 0.6
                        else:
                            surveyors_report += f"; {surrounding_text.strip()}"
                            confidence = min(confidence + 0.1, 0.9)
            
            # Clean up and validate
            if surveyors_report:
                # Remove excessive whitespace
                surveyors_report = ' '.join(surveyors_report.split())
                
                # Truncate if too long
                if len(surveyors_report) > 500:
                    surveyors_report = surveyors_report[:497] + "..."
                
                # Increase confidence for specific indicators
                status_indicators = ['attached', 'available', 'provided', 'completed', 'pending']
                if any(indicator in surveyors_report.lower() for indicator in status_indicators):
                    confidence = min(confidence + 0.1, 1.0)
        
        except Exception as e:
            logger.warning(f"Error extracting surveyor's report: {str(e)}")
            confidence = 0.0
        
        return surveyors_report, confidence
    
    def extract_climate_change_risk(self, text: str) -> Tuple[Optional[str], float]:
        """
        Extract climate change risk assessment details
        
        Args:
            text: Text to extract from
            
        Returns:
            Tuple of (climate risk description, confidence score)
        """
        confidence = 0.0
        climate_risk = None
        
        try:
            # Try pattern matching first
            matches = self.patterns['climate_change_risk_pattern'].findall(text)
            if matches:
                climate_risk = matches[0].strip()
                confidence = 0.8
            
            # Try keyword-based extraction
            if not climate_risk:
                climate_keywords_found = []
                for keyword in self.keywords['climate_change_keywords']:
                    if keyword.lower() in text.lower():
                        climate_keywords_found.append(keyword)
                
                if climate_keywords_found:
                    # Extract surrounding context
                    for keyword in climate_keywords_found[:3]:
                        keyword_pos = text.lower().find(keyword.lower())
                        surrounding_text = text[max(0, keyword_pos-40):keyword_pos+160]
                        
                        if not climate_risk:
                            climate_risk = surrounding_text.strip()
                            confidence = 0.6
                        else:
                            climate_risk += f"; {surrounding_text.strip()}"
                            confidence = min(confidence + 0.1, 0.9)
            
            # Clean up and validate
            if climate_risk:
                # Remove excessive whitespace
                climate_risk = ' '.join(climate_risk.split())
                
                # Truncate if too long
                if len(climate_risk) > 500:
                    climate_risk = climate_risk[:497] + "..."
                
                # Increase confidence for specific risk levels
                risk_levels = ['low', 'medium', 'high', 'significant', 'minimal', 'moderate', 'severe']
                if any(level in climate_risk.lower() for level in risk_levels):
                    confidence = min(confidence + 0.15, 1.0)
                
                # Check for specific climate frameworks
                frameworks = ['tcfd', 'task force', 'scenario analysis', 'physical risk', 'transition risk']
                if any(framework in climate_risk.lower() for framework in frameworks):
                    confidence = min(confidence + 0.1, 1.0)
        
        except Exception as e:
            logger.warning(f"Error extracting climate change risk: {str(e)}")
            confidence = 0.0
        
        return climate_risk, confidence
    
    def extract_esg_risk_assessment(self, text: str) -> Tuple[Optional[str], float]:
        """
        Extract ESG (Environmental, Social, Governance) risk assessment details
        
        Args:
            text: Text to extract from
            
        Returns:
            Tuple of (ESG risk description, confidence score)
        """
        confidence = 0.0
        esg_risk = None
        
        try:
            # Try pattern matching first
            matches = self.patterns['esg_risk_assessment_pattern'].findall(text)
            if matches:
                esg_risk = matches[0].strip()
                confidence = 0.8
            
            # Try keyword-based extraction
            if not esg_risk:
                esg_keywords_found = []
                for keyword in self.keywords['esg_keywords']:
                    if keyword.lower() in text.lower():
                        esg_keywords_found.append(keyword)
                
                if esg_keywords_found:
                    # Extract surrounding context
                    for keyword in esg_keywords_found[:3]:
                        keyword_pos = text.lower().find(keyword.lower())
                        surrounding_text = text[max(0, keyword_pos-40):keyword_pos+160]
                        
                        if not esg_risk:
                            esg_risk = surrounding_text.strip()
                            confidence = 0.6
                        else:
                            esg_risk += f"; {surrounding_text.strip()}"
                            confidence = min(confidence + 0.1, 0.9)
            
            # Clean up and validate
            if esg_risk:
                # Remove excessive whitespace
                esg_risk = ' '.join(esg_risk.split())
                
                # Truncate if too long
                if len(esg_risk) > 500:
                    esg_risk = esg_risk[:497] + "..."
                
                # Increase confidence for specific risk assessments
                risk_indicators = ['assessed', 'evaluated', 'rating', 'score', 'compliance', 'framework']
                if any(indicator in esg_risk.lower() for indicator in risk_indicators):
                    confidence = min(confidence + 0.15, 1.0)
                
                # Check for specific ESG components
                esg_components = ['environmental', 'social', 'governance', 'sustainability', 'diversity']
                component_count = sum(1 for comp in esg_components if comp in esg_risk.lower())
                if component_count >= 2:
                    confidence = min(confidence + 0.1, 1.0)
        
        except Exception as e:
            logger.warning(f"Error extracting ESG risk assessment: {str(e)}")
            confidence = 0.0
        
        return esg_risk, confidence
    
    def generate_excel_report(self, extracted_data_list: List[ExtractedData], output_path: str) -> str:
        """
        Enhanced Excel report generation for Analysis document format with:
        - Multiple RI Slip processing with data merging and gap-filling logic
        - Field completeness tracking and confidence scoring per field (all fields critical)
        - Data validation rules for financial amounts, percentages, and dates
        - Working sheet format output matching Analysis document exactly
        - Processing notes and data source tracking per field
        
        Args:
            extracted_data_list: List of extracted data objects from multiple RI Slips
            output_path: Path for output Excel file
            
        Returns:
            Path to generated Excel file
        """
        try:
            # Step 1: Process multiple RI Slips with data merging and gap-filling
            merged_data = self._merge_multiple_ri_slips(extracted_data_list)
            
            # Step 2: Convert to Analysis document format with enhanced validation
            rows = []
            for data in merged_data:
                # Migrate legacy data to analysis format
                data.migrate_to_analysis_format()
                
                # Validate data according to Analysis document rules
                validation_result = self._validate_analysis_document_data(data.analysis_data)
                
                # Calculate field-level confidence scores
                field_confidence = self._calculate_field_confidence_scores(data)
                data.analysis_data.field_confidence_scores = field_confidence
                
                # Track data completeness for all 23 critical fields
                completeness = self._track_field_completeness(data.analysis_data)
                data.analysis_data.data_completeness = completeness
                
                # Add validation notes to processing notes
                if not validation_result.is_valid:
                    data.analysis_data.processing_notes.extend([
                        f"Validation Error: {error}" for error in validation_result.errors
                    ])
                if validation_result.warnings:
                    data.analysis_data.processing_notes.extend([
                        f"Validation Warning: {warning}" for warning in validation_result.warnings
                    ])
                
                # Use the enhanced Analysis document Excel row format
                row = self._to_enhanced_analysis_excel_row(data)
                rows.append(row)
            
            # Step 3: Create DataFrame with Analysis document columns
            df = pd.DataFrame(rows, columns=self.excel_template['columns'])
            
            # Step 4: Write to Excel with enhanced formatting
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Analysis Working Sheet', index=False)
                
                # Get workbook and worksheet for formatting
                workbook = writer.book
                worksheet = writer.sheets['Analysis Working Sheet']
                
                # Apply enhanced formatting
                self._apply_enhanced_excel_formatting(worksheet, df)
                
                # Add data validation sheets
                self._add_validation_summary_sheet(writer, merged_data)
                self._add_field_completeness_sheet(writer, merged_data)
                self._add_confidence_analysis_sheet(writer, merged_data)
                self._add_processing_notes_sheet(writer, merged_data)
            
            logger.info(f"Enhanced Excel report generated: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error generating enhanced Excel report: {str(e)}")
            raise
    
    def _merge_multiple_ri_slips(self, extracted_data_list: List[ExtractedData]) -> List[ExtractedData]:
        """
        Merge multiple RI Slip data with gap-filling logic
        
        Args:
            extracted_data_list: List of extracted data from multiple RI Slips
            
        Returns:
            List of merged extracted data with gap-filled information
        """
        if not extracted_data_list:
            return []
        
        # Group by reference number or similar identifier
        grouped_data = {}
        
        for data in extracted_data_list:
            # Use reference number as primary key, fallback to insured name
            key = (data.analysis_data.reference_number or 
                   data.analysis_data.insured_name or 
                   f"unknown_{len(grouped_data)}")
            
            if key not in grouped_data:
                grouped_data[key] = []
            grouped_data[key].append(data)
        
        # Merge data for each group
        merged_results = []
        for key, data_group in grouped_data.items():
            if len(data_group) == 1:
                # Single document, no merging needed
                merged_results.append(data_group[0])
            else:
                # Multiple documents, merge with gap-filling
                merged_data = self._merge_data_group(data_group)
                merged_results.append(merged_data)
        
        return merged_results
    
    def _merge_data_group(self, data_group: List[ExtractedData]) -> ExtractedData:
        """
        Merge a group of ExtractedData objects with gap-filling logic
        
        Args:
            data_group: List of ExtractedData objects to merge
            
        Returns:
            Merged ExtractedData object
        """
        # Start with the first data object as base
        merged = data_group[0]
        
        # Merge analysis data from all documents
        for data in data_group[1:]:
            merged = self._fill_gaps_in_analysis_data(merged, data)
        
        # Update source documents list
        source_docs = []
        for data in data_group:
            source_docs.extend(data.analysis_data.source_documents)
        merged.analysis_data.source_documents = list(set(source_docs))
        
        # Update processing notes
        processing_notes = []
        for data in data_group:
            processing_notes.extend(data.analysis_data.processing_notes)
        processing_notes.append(f"Merged data from {len(data_group)} documents")
        merged.analysis_data.processing_notes = processing_notes
        
        # Calculate average confidence score
        confidence_scores = [data.confidence_score for data in data_group if data.confidence_score > 0]
        if confidence_scores:
            merged.confidence_score = sum(confidence_scores) / len(confidence_scores)
            merged.analysis_data.confidence_score = merged.confidence_score
        
        return merged
    
    def _fill_gaps_in_analysis_data(self, base_data: ExtractedData, fill_data: ExtractedData) -> ExtractedData:
        """
        Fill gaps in base_data using fill_data
        
        Args:
            base_data: Base ExtractedData object
            fill_data: ExtractedData object to use for filling gaps
            
        Returns:
            ExtractedData object with filled gaps
        """
        # Get all field names from AnalysisDocumentData
        field_names = [
            'reference_number', 'date_received', 'insured_name', 'cedant_reinsured', 'broker_name',
            'perils_covered', 'geographical_limit', 'situation_of_risk', 'occupation_of_insured', 
            'main_activities', 'total_sums_insured', 'currency', 'excess_retention', 'premium_rates',
            'period_of_insurance', 'pml_percentage', 'cat_exposure', 'reinsurance_deductions',
            'claims_experience_3_years', 'share_offered_percentage', 'surveyors_report',
            'climate_change_risk', 'esg_risk_assessment'
        ]
        
        # Fill gaps where base_data has None/empty values
        for field_name in field_names:
            base_value = getattr(base_data.analysis_data, field_name, None)
            fill_value = getattr(fill_data.analysis_data, field_name, None)
            
            # Fill if base is empty and fill has value
            if (not base_value or base_value == "") and fill_value:
                setattr(base_data.analysis_data, field_name, fill_value)
                
                # Track the source of filled data
                if field_name not in base_data.analysis_data.field_confidence_scores:
                    base_data.analysis_data.field_confidence_scores[field_name] = 0.0
                
                # Update confidence based on fill source
                fill_confidence = getattr(fill_data.analysis_data.field_confidence_scores, field_name, 0.5)
                base_data.analysis_data.field_confidence_scores[field_name] = max(
                    base_data.analysis_data.field_confidence_scores[field_name],
                    fill_confidence * 0.8  # Slightly reduce confidence for filled data
                )
        
        return base_data
    
    def _validate_analysis_document_data(self, analysis_data: AnalysisDocumentData) -> ValidationResult:
        """
        Validate Analysis document data according to business rules
        
        Args:
            analysis_data: AnalysisDocumentData object to validate
            
        Returns:
            ValidationResult with errors and warnings
        """
        
        errors = []
        warnings = []
        
        # Validate financial amounts (must be non-negative)
        financial_fields = ['total_sums_insured', 'excess_retention', 'reinsurance_deductions']
        for field in financial_fields:
            value = getattr(analysis_data, field, None)
            if value is not None and value < 0:
                errors.append(f"{field} cannot be negative: {value}")
        
        # Validate percentages (must be 0-100)
        percentage_fields = ['premium_rates', 'pml_percentage', 'share_offered_percentage']
        for field in percentage_fields:
            value = getattr(analysis_data, field, None)
            if value is not None and (value < 0 or value > 100):
                errors.append(f"{field} must be between 0 and 100: {value}")
        
        # Validate currency code (must be 3 characters)
        if analysis_data.currency and len(analysis_data.currency) != 3:
            errors.append(f"Currency code must be 3 characters: {analysis_data.currency}")
        
        # Validate dates
        if analysis_data.date_received and analysis_data.date_received > datetime.now():
            warnings.append(f"Date received is in the future: {analysis_data.date_received}")
        
        # Check for critical missing fields
        missing_critical = analysis_data.get_missing_critical_fields()
        if missing_critical:
            warnings.extend([f"Missing critical field: {field}" for field in missing_critical])
        
        # Validate reference number format (basic check)
        if analysis_data.reference_number and len(analysis_data.reference_number) < 3:
            warnings.append(f"Reference number seems too short: {analysis_data.reference_number}")
        
        return ValidationResult(
            is_valid=len(errors) == 0,
            errors=errors,
            warnings=warnings
        )
    
    def _calculate_field_confidence_scores(self, data: ExtractedData) -> Dict[str, float]:
        """
        Calculate confidence scores for each field in the Analysis document
        
        Args:
            data: ExtractedData object
            
        Returns:
            Dictionary mapping field names to confidence scores (0.0-1.0)
        """
        field_confidence = {}
        
        # Base confidence from overall extraction
        base_confidence = data.confidence_score
        
        # Field-specific confidence adjustments
        field_adjustments = {
            # High confidence fields (commonly available)
            'reference_number': 0.9,
            'insured_name': 0.9,
            'cedant_reinsured': 0.9,
            'broker_name': 0.9,
            'total_sums_insured': 0.8,
            'currency': 0.8,
            'perils_covered': 0.8,
            'situation_of_risk': 0.8,
            'period_of_insurance': 0.8,
            
            # Medium confidence fields
            'date_received': 0.7,
            'geographical_limit': 0.7,
            'occupation_of_insured': 0.7,
            'main_activities': 0.7,
            'excess_retention': 0.7,
            
            # Lower confidence fields (may need analysis)
            'premium_rates': 0.6,
            'pml_percentage': 0.5,
            'cat_exposure': 0.5,
            'reinsurance_deductions': 0.5,
            'claims_experience_3_years': 0.5,
            'share_offered_percentage': 0.6,
            'surveyors_report': 0.4,
            'climate_change_risk': 0.3,
            'esg_risk_assessment': 0.3
        }
        
        # Calculate confidence for each field
        for field_name, adjustment in field_adjustments.items():
            field_value = getattr(data.analysis_data, field_name, None)
            
            if field_value is not None and field_value != "":
                # Field has value, calculate confidence
                field_confidence[field_name] = min(base_confidence * adjustment, 1.0)
            else:
                # Field is missing
                field_confidence[field_name] = 0.0
        
        return field_confidence
    
    def _track_field_completeness(self, analysis_data: AnalysisDocumentData) -> Dict[str, bool]:
        """
        Track completeness for all 23 critical fields
        
        Args:
            analysis_data: AnalysisDocumentData object
            
        Returns:
            Dictionary mapping field names to completeness status
        """
        completeness = {}
        
        # All 23 critical fields
        critical_fields = [
            'reference_number', 'date_received', 'insured_name', 'cedant_reinsured', 'broker_name',
            'perils_covered', 'geographical_limit', 'situation_of_risk', 'occupation_of_insured',
            'main_activities', 'total_sums_insured', 'currency', 'excess_retention', 'premium_rates',
            'period_of_insurance', 'pml_percentage', 'cat_exposure', 'reinsurance_deductions',
            'claims_experience_3_years', 'share_offered_percentage', 'surveyors_report',
            'climate_change_risk', 'esg_risk_assessment'
        ]
        
        for field_name in critical_fields:
            field_value = getattr(analysis_data, field_name, None)
            completeness[field_name] = field_value is not None and field_value != ""
        
        return completeness
    
    def _to_enhanced_analysis_excel_row(self, data: ExtractedData) -> Dict[str, Any]:
        """
        Convert ExtractedData to enhanced Analysis document Excel row format
        
        Args:
            data: ExtractedData object
            
        Returns:
            Dictionary representing Excel row with enhanced formatting
        """
        row_data = {}
        
        # Field mapping from Analysis document to Excel columns
        field_mapping = {
            'Reference Number': 'reference_number',
            'Date Received': 'date_received',
            'Insured': 'insured_name',
            'Cedant/Reinsured': 'cedant_reinsured',
            'Broker': 'broker_name',
            'Perils Covered': 'perils_covered',
            'Geographical Limit': 'geographical_limit',
            'Situation of Risk/Voyage': 'situation_of_risk',
            'Occupation of Insured': 'occupation_of_insured',
            'Main Activities': 'main_activities',
            'Total Sums Insured': 'total_sums_insured',
            'Currency': 'currency',
            'Excess/Retention': 'excess_retention',
            'Premium Rates (%)': 'premium_rates',
            'Period of Insurance': 'period_of_insurance',
            'PML %': 'pml_percentage',
            'CAT Exposure': 'cat_exposure',
            'Reinsurance Deductions': 'reinsurance_deductions',
            'Claims Experience (3 years)': 'claims_experience_3_years',
            'Share offered %': 'share_offered_percentage',
            'Surveyor\'s Report': 'surveyors_report',
            'Climate Change Risk': 'climate_change_risk',
            'ESG Risk Assessment': 'esg_risk_assessment'
        }
        
        # Populate data fields with enhanced formatting
        for excel_col, field_name in field_mapping.items():
            value = getattr(data.analysis_data, field_name, None)
            
            # Format values appropriately for Excel
            if isinstance(value, Decimal):
                row_data[excel_col] = float(value)
            elif isinstance(value, datetime):
                row_data[excel_col] = value.strftime('%Y-%m-%d')
            elif value is None:
                row_data[excel_col] = ""
            else:
                row_data[excel_col] = str(value)
        
        # Add enhanced metadata columns
        row_data['Confidence Score'] = f"{data.analysis_data.confidence_score:.2%}"
        row_data['Data Completeness %'] = f"{data.analysis_data.calculate_completeness_score():.1%}"
        row_data['Processing Notes'] = '; '.join(data.analysis_data.processing_notes) if data.analysis_data.processing_notes else ""
        row_data['Source Documents'] = '; '.join(data.analysis_data.source_documents) if data.analysis_data.source_documents else ""
        
        return row_data
    
    def _apply_enhanced_excel_formatting(self, worksheet, df):
        """
        Apply enhanced formatting to the Excel worksheet
        
        Args:
            worksheet: openpyxl worksheet object
            df: pandas DataFrame with the data
        """
        from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
        from openpyxl.formatting.rule import ColorScaleRule
        
        # Define colors and styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Color coding for field types
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Commonly available
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # May need analysis
        red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")    # Missing/low confidence
        
        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header formatting
        for col_idx, col_name in enumerate(self.excel_template['columns'], 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Color-code based on field availability
            if col_name in self.excel_template['commonly_available']:
                cell.fill = green_fill
                cell.font = Font(bold=True)
            elif col_name in self.excel_template['may_need_analysis']:
                cell.fill = yellow_fill
                cell.font = Font(bold=True)
        
        # Apply data formatting
        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, len(self.excel_template['columns']) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Auto-adjust column widths with better sizing
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set appropriate width based on content
            if max_length < 10:
                adjusted_width = 12
            elif max_length < 20:
                adjusted_width = max_length + 2
            elif max_length < 50:
                adjusted_width = max_length + 5
            else:
                adjusted_width = 50
            
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add conditional formatting for confidence scores
        confidence_col = None
        for col_idx, col_name in enumerate(self.excel_template['columns'], 1):
            if col_name == 'Confidence Score':
                confidence_col = col_idx
                break
        
        if confidence_col:
            # Apply color scale to confidence scores
            color_scale_rule = ColorScaleRule(
                start_type='num', start_value=0, start_color='FF0000',  # Red for low
                mid_type='num', mid_value=0.5, mid_color='FFFF00',     # Yellow for medium
                end_type='num', end_value=1, end_color='00FF00'        # Green for high
            )
            
            col_letter = worksheet.cell(row=1, column=confidence_col).column_letter
            worksheet.conditional_formatting.add(
                f'{col_letter}2:{col_letter}{len(df) + 1}',
                color_scale_rule
            )
    
    def _add_validation_summary_sheet(self, writer, merged_data: List[ExtractedData]):
        """Add validation summary sheet to Excel workbook"""
        validation_data = []
        
        for idx, data in enumerate(merged_data, 1):
            validation_result = self._validate_analysis_document_data(data.analysis_data)
            
            validation_data.append({
                'Row': idx,
                'Reference': data.analysis_data.reference_number or f"Row {idx}",
                'Valid': 'Yes' if validation_result.is_valid else 'No',
                'Error Count': len(validation_result.errors),
                'Warning Count': len(validation_result.warnings),
                'Errors': '; '.join(validation_result.errors),
                'Warnings': '; '.join(validation_result.warnings)
            })
        
        validation_df = pd.DataFrame(validation_data)
        validation_df.to_excel(writer, sheet_name='Validation Summary', index=False)
    
    def _add_field_completeness_sheet(self, writer, merged_data: List[ExtractedData]):
        """Add field completeness analysis sheet to Excel workbook"""
        if not merged_data:
            return
        
        # Get all field names
        field_names = list(merged_data[0].analysis_data.data_completeness.keys())
        
        completeness_data = []
        for field_name in field_names:
            completed_count = sum(1 for data in merged_data 
                                if data.analysis_data.data_completeness.get(field_name, False))
            total_count = len(merged_data)
            completeness_percentage = (completed_count / total_count * 100) if total_count > 0 else 0
            
            completeness_data.append({
                'Field Name': field_name,
                'Completed Count': completed_count,
                'Total Count': total_count,
                'Completeness %': f"{completeness_percentage:.1f}%",
                'Field Type': self._get_field_type(field_name)
            })
        
        completeness_df = pd.DataFrame(completeness_data)
        completeness_df.to_excel(writer, sheet_name='Field Completeness', index=False)
    
    def _add_confidence_analysis_sheet(self, writer, merged_data: List[ExtractedData]):
        """Add confidence analysis sheet to Excel workbook"""
        if not merged_data:
            return
        
        confidence_data = []
        
        # Get all unique field names from confidence scores
        all_fields = set()
        for data in merged_data:
            all_fields.update(data.analysis_data.field_confidence_scores.keys())
        
        for field_name in sorted(all_fields):
            confidence_scores = [
                data.analysis_data.field_confidence_scores.get(field_name, 0.0)
                for data in merged_data
            ]
            
            if confidence_scores:
                avg_confidence = sum(confidence_scores) / len(confidence_scores)
                min_confidence = min(confidence_scores)
                max_confidence = max(confidence_scores)
                
                confidence_data.append({
                    'Field Name': field_name,
                    'Average Confidence': f"{avg_confidence:.2%}",
                    'Min Confidence': f"{min_confidence:.2%}",
                    'Max Confidence': f"{max_confidence:.2%}",
                    'Records with Data': sum(1 for score in confidence_scores if score > 0),
                    'Total Records': len(confidence_scores)
                })
        
        confidence_df = pd.DataFrame(confidence_data)
        confidence_df.to_excel(writer, sheet_name='Confidence Analysis', index=False)
    
    def _add_processing_notes_sheet(self, writer, merged_data: List[ExtractedData]):
        """Add processing notes sheet to Excel workbook"""
        notes_data = []
        
        for idx, data in enumerate(merged_data, 1):
            reference = data.analysis_data.reference_number or f"Row {idx}"
            
            if data.analysis_data.processing_notes:
                for note in data.analysis_data.processing_notes:
                    notes_data.append({
                        'Row': idx,
                        'Reference': reference,
                        'Note Type': self._classify_note_type(note),
                        'Processing Note': note,
                        'Source Documents': '; '.join(data.analysis_data.source_documents)
                    })
            else:
                notes_data.append({
                    'Row': idx,
                    'Reference': reference,
                    'Note Type': 'Info',
                    'Processing Note': 'No processing notes',
                    'Source Documents': '; '.join(data.analysis_data.source_documents)
                })
        
        notes_df = pd.DataFrame(notes_data)
        notes_df.to_excel(writer, sheet_name='Processing Notes', index=False)
    
    def _get_field_type(self, field_name: str) -> str:
        """Get the type category for a field"""
        if field_name in self.excel_template['commonly_available']:
            return 'Commonly Available'
        elif field_name in self.excel_template['may_need_analysis']:
            return 'May Need Analysis'
        else:
            return 'Other'
    
    def _classify_note_type(self, note: str) -> str:
        """Classify processing note type"""
        note_lower = note.lower()
        if 'error' in note_lower:
            return 'Error'
        elif 'warning' in note_lower:
            return 'Warning'
        elif 'validation' in note_lower:
            return 'Validation'
        elif 'merged' in note_lower:
            return 'Merge'
        else:
            return 'Info'

    def process_documents(self, file_paths: List[str]) -> List[ExtractedData]:
        """
        Process multiple documents and extract structured data
        
        Args:
            file_paths: List of document file paths
            
        Returns:
            List of extracted data objects
        """
        extracted_data_list = []
        
        for file_path in file_paths:
            try:
                logger.info(f"Processing document: {file_path}")
                
                # Process document with OCR agent
                result = ocr_agent.process_document(file_path)
                
                if isinstance(result, EmailContent):
                    # Extract from email and attachments
                    extracted_data = self.extract_from_email(result)
                    extracted_data_list.append(extracted_data)
                
                elif isinstance(result, OCRResult):
                    # Extract from OCR result
                    extracted_data = ExtractedData()
                    self._extract_from_text(result.text, extracted_data)
                    extracted_data.confidence_score = self._calculate_confidence(extracted_data)
                    extracted_data_list.append(extracted_data)
                
                elif isinstance(result, (ExcelData, WordDocumentData, PowerPointData)):
                    # Extract from structured document
                    extracted_data = ExtractedData()
                    
                    if isinstance(result, WordDocumentData):
                        self._extract_from_text(result.text, extracted_data)
                        for table in result.tables:
                            self._extract_from_table(table, extracted_data)
                    
                    elif isinstance(result, ExcelData):
                        for sheet_data in result.sheets.values():
                            self._extract_from_excel_sheet(sheet_data, extracted_data)
                    
                    elif isinstance(result, PowerPointData):
                        self._extract_from_text(result.text, extracted_data)
                    
                    extracted_data.confidence_score = self._calculate_confidence(extracted_data)
                    extracted_data_list.append(extracted_data)
                
            except Exception as e:
                logger.error(f"Error processing document {file_path}: {str(e)}")
                # Create error entry
                error_data = ExtractedData()
                error_data.processing_notes.append(f"Processing error: {str(e)}")
                error_data.confidence_score = 0.0
                extracted_data_list.append(error_data)
        
        return extracted_data_list


# Global data extraction agent instance
data_extraction_agent = DataExtractionAgent()